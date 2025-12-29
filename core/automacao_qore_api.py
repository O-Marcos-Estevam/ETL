"""
================================================================================
AUTOMACAO QORE API V1 - BASEADA EM HTTP REQUESTS
================================================================================

Substitui Selenium por chamadas HTTP diretas (requests).
Muito mais rapido e eficiente que a versao com browser.

Vantagens:
- ~10x mais rapido (30s vs 5min para 34 fundos)
- Baixo consumo de memoria (~50MB vs ~500MB/Chrome)
- Downloads paralelos com ThreadPoolExecutor
- Sem dependencia de Chrome/Driver
- Mais robusto (sem fragilidade de DOM)

Endpoints utilizados:
- POST /api/v1/authorize -> JWT Bearer token
- GET /api/v1/fundos-posicao/{uuid}/arquivos?tipo=X&p=0 -> Lista arquivos
- Download direto via URL do arquivo

Autor: ETL Team
Data: Dezembro 2025
================================================================================
"""

import os
import re
import sys
import json
import time
import shutil
import logging
import zipfile
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Any
from concurrent.futures import ThreadPoolExecutor, as_completed
from enum import Enum
from urllib.parse import urljoin
import threading

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import pandas as pd
import openpyxl


# =============================================================================
# CONFIGURACAO DE LOGGING
# =============================================================================

class LogFormatter(logging.Formatter):
    """Formatter customizado com cores e simbolos."""

    COLORS = {
        'DEBUG': '\033[36m',     # Cyan
        'INFO': '\033[32m',      # Green
        'WARNING': '\033[33m',   # Yellow
        'ERROR': '\033[31m',     # Red
        'CRITICAL': '\033[35m',  # Magenta
        'RESET': '\033[0m'
    }

    SYMBOLS = {
        'DEBUG': '[.]',
        'INFO': '[+]',
        'WARNING': '[!]',
        'ERROR': '[X]',
        'CRITICAL': '[!!!]'
    }

    def format(self, record):
        symbol = self.SYMBOLS.get(record.levelname, '[?]')
        color = self.COLORS.get(record.levelname, '')
        reset = self.COLORS['RESET']
        record.msg = f"{color}{symbol}{reset} {record.msg}"
        return super().format(record)


def setup_logging(level=logging.INFO) -> logging.Logger:
    """Configura e retorna o logger principal."""
    logger = logging.getLogger('QoreAPI')
    logger.setLevel(level)

    if not logger.handlers:
        handler = logging.StreamHandler(sys.stdout)
        handler.setFormatter(LogFormatter('%(message)s'))
        logger.addHandler(handler)

    return logger


log = setup_logging()


# =============================================================================
# CONSTANTES E CONFIGURACOES
# =============================================================================

class FileType(Enum):
    """Tipos de arquivo suportados."""
    PDF = 'PDF'
    EXCEL = 'EXCEL'
    XML = 'XML'


@dataclass
class APIConfig:
    """Configuracoes da API."""
    BASE_URL: str = 'https://hub.qoredtvm.com.br'
    AUTH_ENDPOINT: str = '/api/v1/authorize'
    FILES_ENDPOINT: str = '/api/v1/fundos-posicao/{uuid}/arquivos'
    FUND_ENDPOINT: str = '/api/v1/fundos-posicao/{uuid}'
    DOWNLOAD_ENDPOINT: str = '/api/v1/fundos-posicao/{fundo_uuid}/arquivos/{arquivo_guid}/download'

    # Mapeamento de tipo para parametro da API
    # Descoberto via captura_xml_endpoint.py usando Chrome DevTools Protocol
    TYPE_PARAMS: Dict[str, str] = field(default_factory=lambda: {
        'PDF': 'CARTEIRA_PDF',
        'EXCEL': 'CARTEIRA_EXCEL',
        'XML': 'XML_5_0'  # XML Anbima 5.0 - Confirmado!
    })

    # Timeouts
    CONNECT_TIMEOUT: int = 10
    READ_TIMEOUT: int = 60
    MAX_RETRIES: int = 3
    BACKOFF_FACTOR: float = 0.5

    # Workers
    NUM_WORKERS: int = 10


@dataclass
class ReportConfig:
    """Configuracao para cada tipo de relatorio."""
    extension: str
    type_name: str
    api_param: Optional[str]  # None se nao disponivel via API

    def get_filename(self, fundo: str, data: datetime) -> str:
        """Gera o nome base do arquivo."""
        return f"{data.strftime('%d.%m')} - {self.type_name} - {fundo}"


# Configuracoes dos tipos de relatorio
# Parametros descobertos via captura_xml_endpoint.py (Chrome DevTools Protocol)
REPORT_CONFIGS: Dict[str, ReportConfig] = {
    'PDF': ReportConfig('.pdf', 'Carteira Diaria', 'CARTEIRA_PDF'),
    'EXCEL': ReportConfig('.xlsx', 'Carteira Excel', 'CARTEIRA_EXCEL'),
    'XML': ReportConfig('.xml', 'Carteira XML', 'XML_5_0')  # XML Anbima 5.0 - Confirmado!
}

# Mapeamento de fundos BLOKO para padroes de busca em arquivos
BLOKO_PATTERNS: Dict[str, str] = {
    'BLOKO URBANISMO': 'urbanismo',
    'BLOKO FIM': 'fundo-de-investimento'
}

# Meses por extenso
MESES_EXTENSO: Dict[str, str] = {
    '01': 'Janeiro', '02': 'Fevereiro', '03': 'Marco', '04': 'Abril',
    '05': 'Maio', '06': 'Junho', '07': 'Julho', '08': 'Agosto',
    '09': 'Setembro', '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro'
}


# =============================================================================
# DATACLASSES DE CONFIGURACAO (Reutilizadas do automacao_qore.py)
# =============================================================================

@dataclass
class QoreCredentials:
    """Credenciais de acesso ao QORE."""
    url: str
    email: str
    senha: str


@dataclass
class QorePaths:
    """Caminhos de destino dos arquivos."""
    pdf: str = ''
    excel: str = ''
    xml: str = ''
    pdf_monitoramento: str = ''
    temp_download: str = ''
    bd_path: str = ''
    base_fundos: str = r'C:\bloko\Fundos - Documentos'


@dataclass
class QoreFlags:
    """Flags de controle da automacao."""
    qore_enabled: bool = False
    pdf_enabled: bool = False
    pdf_lote: bool = False
    excel_enabled: bool = False
    excel_lote: bool = False
    xml_enabled: bool = False
    xml_lote: bool = False


@dataclass
class QoreDatas:
    """Datas de referencia."""
    data_inicial: datetime = None
    data_final: datetime = None

    @property
    def data_exibicao(self) -> str:
        """Data formatada para exibicao (dd/mm/yyyy)."""
        return self.data_inicial.strftime('%d/%m/%Y') if self.data_inicial else ''

    @property
    def is_lote(self) -> bool:
        """Verifica se e modo lote (datas diferentes)."""
        return self.data_inicial != self.data_final


# =============================================================================
# FUNCOES UTILITARIAS
# =============================================================================

def validar_boolean(valor) -> bool:
    """Converte valores da planilha para booleano."""
    if pd.isna(valor) or valor is None:
        return False
    valor_str = str(valor).strip().upper()
    return valor_str in {'SIM', 'S', 'TRUE', 'VERDADEIRO', 'YES', 'Y', '1'}


def extrair_data_de_nome_arquivo(nome_arquivo: str) -> Optional[datetime]:
    """Extrai data do nome do arquivo (formato YYYYMMDD)."""
    matches = re.findall(r'_(\d{8})', nome_arquivo)
    if not matches:
        matches = re.findall(r'(\d{8})', nome_arquivo)

    for date_str in matches:
        try:
            dt = datetime.strptime(date_str, '%Y%m%d')
            if 2000 <= dt.year <= 2035:
                return dt
        except ValueError:
            continue
    return None


def get_versioned_filepath(directory: str, base_name: str, extension: str) -> str:
    """Gera caminho com versionamento automatico."""
    target = os.path.join(directory, f"{base_name}{extension}")

    if not os.path.exists(target):
        return target

    version = 1
    while True:
        versioned = os.path.join(directory, f"{base_name} ({version}){extension}")
        if not os.path.exists(versioned):
            return versioned
        version += 1


# =============================================================================
# CLASSE: FundoManager (Reutilizada)
# =============================================================================

class FundoManager:
    """Gerencia a lista de fundos e suas siglas."""

    def __init__(self, bd_path: str):
        self.bd_path = bd_path
        self.fundos: Dict[str, str] = {}  # {nome_fundo: pasta_destino}
        self.siglas: Dict[str, str] = {}  # {nome_fundo: sigla_busca}

    def carregar_fundos(self) -> bool:
        """Carrega fundos do arquivo BD.xlsx."""
        try:
            df = pd.read_excel(
                self.bd_path,
                sheet_name='BD',
                engine='openpyxl',
                header=None,
                usecols=[1, 2, 9],
                dtype=str
            )

            for _, row in df.iterrows():
                flag_qore = str(row.iloc[2] or '').strip().upper()

                if flag_qore in {'SIM', 'S', 'TRUE', 'YES', 'VERDADEIRO', 'QORE'}:
                    apelido = row.iloc[0]
                    caminho = row.iloc[1]

                    if pd.notna(apelido) and pd.notna(caminho):
                        apelido_clean = str(apelido).strip()
                        caminho_clean = str(caminho).strip()

                        if apelido_clean and caminho_clean:
                            nome_final = self._processar_nome_bloko(apelido_clean)
                            self.fundos[nome_final] = caminho_clean

            self._gerar_siglas()
            log.info(f"Carregados {len(self.fundos)} fundos QORE")
            return len(self.fundos) > 0

        except Exception as e:
            log.error(f"Falha ao ler BD.xlsx: {e}")
            return False

    def _processar_nome_bloko(self, apelido: str) -> str:
        """Processa nomes especiais de fundos BLOKO."""
        partes = apelido.split()
        if len(partes) > 1 and partes[1] == 'BLOKO':
            if partes[0] == 'FIP':
                return 'BLOKO URBANISMO'
            else:
                return 'BLOKO FIM'
        return apelido

    def _gerar_siglas(self):
        """Gera mapa de siglas para busca."""
        for nome in self.fundos:
            if 'BLOKO' in nome.upper():
                self.siglas[nome] = nome
            else:
                partes = nome.split(' ', 1)
                self.siglas[nome] = partes[1].strip() if len(partes) > 1 else nome

    def get_sigla(self, nome_fundo: str) -> str:
        """Retorna a sigla de busca para um fundo."""
        return self.siglas.get(nome_fundo, nome_fundo)

    def is_bloko(self, nome_fundo: str) -> bool:
        """Verifica se e um fundo BLOKO."""
        return 'BLOKO' in nome_fundo.upper()

    def get_bloko_pattern(self, nome_fundo: str) -> str:
        """Retorna o padrao de busca em arquivos para fundos BLOKO."""
        return BLOKO_PATTERNS.get(nome_fundo, '')


# =============================================================================
# CLASSE: FileHandler (Adaptada para API)
# =============================================================================

class FileHandler:
    """Gerencia download e movimentacao de arquivos."""

    def __init__(self, temp_path: str, paths: QorePaths):
        self.temp_path = temp_path
        self.paths = paths
        self.lock = threading.Lock()

    def limpar_temp(self):
        """Limpa pasta temporaria de downloads."""
        if not os.path.exists(self.temp_path):
            os.makedirs(self.temp_path, exist_ok=True)
            return

        for item in os.listdir(self.temp_path):
            item_path = os.path.join(self.temp_path, item)
            try:
                if os.path.isfile(item_path):
                    os.unlink(item_path)
                elif os.path.isdir(item_path):
                    shutil.rmtree(item_path)
            except Exception:
                pass

    def processar_zip(self, zip_path: Path, fundo_nome: str, fundo_manager: FundoManager,
                     report_type: str, data_referencia: datetime) -> int:
        """Processa arquivo ZIP baixado."""
        config = REPORT_CONFIGS.get(report_type.upper())
        if not config:
            return 0

        temp_extract = Path(self.temp_path) / f'extract_{zip_path.stem}'
        temp_extract.mkdir(exist_ok=True)

        try:
            with zipfile.ZipFile(zip_path, 'r') as zf:
                zf.extractall(temp_extract)

            zip_path.unlink()

            arquivos_processados = 0

            if fundo_manager.is_bloko(fundo_nome):
                padrao = fundo_manager.get_bloko_pattern(fundo_nome)
            else:
                padrao = fundo_manager.get_sigla(fundo_nome).lower()

            for arquivo in temp_extract.glob(f'*{config.extension}'):
                if padrao and padrao.lower() not in arquivo.name.lower():
                    continue

                data_arquivo = extrair_data_de_nome_arquivo(arquivo.name)
                if not data_arquivo:
                    data_arquivo = data_referencia

                if self._mover_arquivo(
                    arquivo, fundo_nome, data_arquivo, report_type,
                    fundo_manager.fundos.get(fundo_nome, '')
                ):
                    arquivos_processados += 1

            return arquivos_processados

        finally:
            if temp_extract.exists():
                shutil.rmtree(temp_extract, ignore_errors=True)

    def processar_arquivo_individual(self, arquivo: Path, fundo_nome: str,
                                    data: datetime, report_type: str,
                                    pasta_fundo: str) -> bool:
        """Processa um arquivo individual."""
        return self._mover_arquivo(arquivo, fundo_nome, data, report_type, pasta_fundo)

    def _mover_arquivo(self, arquivo: Path, fundo_nome: str, data: datetime,
                      report_type: str, pasta_fundo: str) -> bool:
        """Move arquivo para destino(s) final(is)."""
        config = REPORT_CONFIGS.get(report_type.upper())
        if not config:
            return False

        try:
            nome_base = config.get_filename(fundo_nome, data)
            destinos = self._get_destinos(report_type, data, pasta_fundo)

            primeiro_destino = None

            with self.lock:
                for i, destino_dir in enumerate(destinos):
                    os.makedirs(destino_dir, exist_ok=True)

                    caminho_final = get_versioned_filepath(
                        destino_dir, nome_base, config.extension
                    )

                    if i == 0:
                        shutil.move(str(arquivo), caminho_final)
                        primeiro_destino = caminho_final
                        log.info(f"  {report_type} -> {Path(caminho_final).name}")
                    else:
                        if primeiro_destino and os.path.exists(primeiro_destino):
                            shutil.copy(primeiro_destino, caminho_final)

            return True

        except Exception as e:
            log.error(f"Falha ao mover arquivo: {e}")
            return False

    def _get_destinos(self, report_type: str, data: datetime, pasta_fundo: str) -> List[str]:
        """Retorna lista de diretorios de destino."""
        destinos = []
        ano = data.strftime('%Y')
        mes = data.strftime('%m')

        if report_type.upper() == 'PDF':
            if pasta_fundo:
                mes_formatado = f"{mes} - {MESES_EXTENSO.get(mes, '')}"
                path_fundo = os.path.join(
                    self.paths.base_fundos, pasta_fundo,
                    '06. Carteiras', ano, mes_formatado
                )
                destinos.append(path_fundo)

            if self.paths.pdf_monitoramento:
                destinos.append(self.paths.pdf_monitoramento)

        elif report_type.upper() == 'EXCEL':
            if self.paths.excel:
                destinos.append(self.paths.excel)

        elif report_type.upper() == 'XML':
            if self.paths.xml:
                destinos.append(self.paths.xml)

        return destinos


# =============================================================================
# CLASSE: QoreAPIClient
# =============================================================================

class QoreAPIClient:
    """Cliente HTTP para API QORE."""

    def __init__(self, credentials: QoreCredentials):
        self.credentials = credentials
        self.config = APIConfig()
        self.token: Optional[str] = None
        self.token_expiry: Optional[datetime] = None
        self.session = self._create_session()

    def _create_session(self) -> requests.Session:
        """Cria sessao HTTP com retry automatico."""
        session = requests.Session()

        retry_strategy = Retry(
            total=self.config.MAX_RETRIES,
            backoff_factor=self.config.BACKOFF_FACTOR,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["HEAD", "GET", "POST", "OPTIONS"]
        )

        adapter = HTTPAdapter(max_retries=retry_strategy)
        session.mount("https://", adapter)
        session.mount("http://", adapter)

        # Headers padroes
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'application/json, text/plain, */*',
            'Content-Type': 'application/json'
        })

        return session

    def authenticate(self) -> bool:
        """Autentica na API e obtem JWT token."""
        try:
            url = urljoin(self.config.BASE_URL, self.config.AUTH_ENDPOINT)

            payload = {
                'username': self.credentials.email,
                'password': self.credentials.senha
            }

            log.info("Autenticando na API QORE...")

            response = self.session.post(
                url,
                json=payload,
                timeout=(self.config.CONNECT_TIMEOUT, self.config.READ_TIMEOUT)
            )

            if response.status_code == 200:
                data = response.json()
                self.token = data.get('access_token') or data.get('token') or data.get('accessToken')

                if self.token:
                    self.session.headers['Authorization'] = f'Bearer {self.token}'
                    log.info("Autenticacao OK!")
                    return True
                else:
                    log.error("Token nao encontrado na resposta")
                    return False
            else:
                log.error(f"Falha na autenticacao: {response.status_code}")
                return False

        except Exception as e:
            log.error(f"Erro na autenticacao: {e}")
            return False

    def get_fund_files(self, uuid: str, file_type: str, page: int = 0) -> List[Dict]:
        """Lista arquivos de um fundo."""
        try:
            endpoint = self.config.FILES_ENDPOINT.format(uuid=uuid)
            url = urljoin(self.config.BASE_URL, endpoint)

            api_param = REPORT_CONFIGS.get(file_type.upper(), REPORT_CONFIGS['PDF']).api_param

            # Verifica se tipo esta disponivel via API
            if api_param is None:
                log.warning(f"{file_type} nao disponivel via API")
                return []

            params = {
                'tipo': api_param,
                'p': page
            }

            response = self.session.get(
                url,
                params=params,
                timeout=(self.config.CONNECT_TIMEOUT, self.config.READ_TIMEOUT)
            )

            if response.status_code == 200:
                data = response.json()
                # A API retorna resposta paginada com 'content'
                if isinstance(data, dict):
                    return data.get('content', data.get('items', data.get('data', [])))
                elif isinstance(data, list):
                    return data
                return []
            else:
                log.warning(f"Falha ao listar arquivos: {response.status_code}")
                return []

        except Exception as e:
            log.error(f"Erro ao listar arquivos: {e}")
            return []

    def download_file_by_guid(self, fundo_uuid: str, arquivo_guid: str, dest_path: Path) -> bool:
        """Baixa um arquivo pelo GUID."""
        try:
            endpoint = self.config.DOWNLOAD_ENDPOINT.format(
                fundo_uuid=fundo_uuid,
                arquivo_guid=arquivo_guid
            )
            url = urljoin(self.config.BASE_URL, endpoint)

            response = self.session.get(
                url,
                timeout=(self.config.CONNECT_TIMEOUT, 120),
                stream=True
            )

            if response.status_code == 200:
                dest_path.parent.mkdir(parents=True, exist_ok=True)

                with open(dest_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                return True
            else:
                log.warning(f"Falha no download: {response.status_code}")
                return False

        except Exception as e:
            log.error(f"Erro no download: {e}")
            return False

    def _request(self, method: str, endpoint: str, **kwargs) -> requests.Response:
        """Wrapper para requisicoes com tratamento de erros."""
        url = urljoin(self.config.BASE_URL, endpoint)

        response = self.session.request(
            method,
            url,
            timeout=(self.config.CONNECT_TIMEOUT, self.config.READ_TIMEOUT),
            **kwargs
        )

        return response


# =============================================================================
# CLASSE: QoreDownloadManager
# =============================================================================

class QoreDownloadManager:
    """Orquestra downloads paralelos."""

    def __init__(self, api_client: QoreAPIClient, file_handler: FileHandler,
                 fundo_manager: FundoManager, datas: QoreDatas):
        self.api = api_client
        self.file_handler = file_handler
        self.fundo_manager = fundo_manager
        self.datas = datas
        self.config = APIConfig()
        self.fundos_uuid: Dict[str, str] = {}  # {nome: uuid}
        self.lock = threading.Lock()

    def load_fundos_uuid(self, json_path: str = None) -> bool:
        """Carrega mapeamento nome -> UUID dos fundos."""
        try:
            # Tenta carregar do arquivo JSON gerado pela captura
            if json_path is None:
                # Caminho padrao relativo ao script
                base_path = Path(__file__).parent.parent
                json_path = base_path / 'dump_qore' / 'fundos' / 'lista_fundos.json'

            if Path(json_path).exists():
                with open(json_path, 'r', encoding='utf-8') as f:
                    fundos_list = json.load(f)

                for fundo in fundos_list:
                    nome = fundo.get('nome', '')
                    url = fundo.get('url', '')

                    # Extrai UUID da URL
                    # URL: https://hub.qoredtvm.com.br/portfolios/fundo-posicao/175a9636-5ff2-...
                    if '/fundo-posicao/' in url:
                        uuid = url.split('/fundo-posicao/')[-1].split('/')[0]

                        # Tenta fazer match com os fundos do BD
                        for nome_bd in self.fundo_manager.fundos:
                            sigla = self.fundo_manager.get_sigla(nome_bd)
                            if sigla.upper() in nome.upper() or nome.upper() in sigla.upper():
                                self.fundos_uuid[nome_bd] = uuid
                                break
                        else:
                            # Se nao encontrou match exato, guarda pelo nome original
                            self.fundos_uuid[nome] = uuid

                log.info(f"Carregados {len(self.fundos_uuid)} UUIDs de fundos")
                return len(self.fundos_uuid) > 0
            else:
                log.warning(f"Arquivo de UUIDs nao encontrado: {json_path}")
                return False

        except Exception as e:
            log.error(f"Erro ao carregar UUIDs: {e}")
            return False

    def download_all_funds(self, file_types: List[str]) -> Dict[str, Dict]:
        """
        Baixa arquivos de todos os fundos em paralelo.
        Processa todos os tipos de arquivo simultaneamente.
        """
        resultados = {}

        # Prepara lista de tarefas
        tasks = []
        for nome_fundo, pasta_fundo in self.fundo_manager.fundos.items():
            uuid = self.fundos_uuid.get(nome_fundo)

            if not uuid:
                # Tenta encontrar por sigla
                sigla = self.fundo_manager.get_sigla(nome_fundo)
                for nome_uuid, uid in self.fundos_uuid.items():
                    if sigla.upper() in nome_uuid.upper():
                        uuid = uid
                        break

            if uuid:
                for file_type in file_types:
                    tasks.append({
                        'nome': nome_fundo,
                        'uuid': uuid,
                        'pasta': pasta_fundo,
                        'tipo': file_type
                    })
            else:
                log.warning(f"UUID nao encontrado para: {nome_fundo}")
                resultados[f"{nome_fundo}_ALL"] = {'status': 'erro', 'motivo': 'UUID nao encontrado'}

        log.info(f"Iniciando {len(tasks)} downloads em paralelo ({self.config.NUM_WORKERS} workers)...")

        # Executa em paralelo
        with ThreadPoolExecutor(max_workers=self.config.NUM_WORKERS) as executor:
            futures = {
                executor.submit(self._download_fund_files, task): task
                for task in tasks
            }

            for future in as_completed(futures):
                task = futures[future]
                try:
                    result = future.result()
                    key = f"{task['nome']}_{task['tipo']}"
                    resultados[key] = result
                except Exception as e:
                    log.error(f"Worker falhou para {task['nome']}: {e}")
                    key = f"{task['nome']}_{task['tipo']}"
                    resultados[key] = {'status': 'erro', 'motivo': str(e)}

        return resultados

    def _download_fund_files(self, task: Dict) -> Dict:
        """Processa download de um fundo/tipo."""
        nome = task['nome']
        uuid = task['uuid']
        pasta = task['pasta']
        tipo = task['tipo']

        resultado = {
            'nome': nome,
            'tipo': tipo,
            'status': 'erro',
            'arquivos': 0
        }

        try:
            # Cria pasta temporaria para este fundo
            temp_fundo = Path(self.file_handler.temp_path) / f'{uuid}_{tipo}'
            temp_fundo.mkdir(parents=True, exist_ok=True)

            # Lista arquivos do fundo
            arquivos = self.api.get_fund_files(uuid, tipo)

            if not arquivos:
                log.warning(f"  {nome} [{tipo}]: Nenhum arquivo encontrado")
                return resultado

            qtd_baixados = 0

            for arq in arquivos:
                # Filtra por data
                data_arq_str = arq.get('data') or arq.get('dataReferencia')
                data_arq = None

                if data_arq_str:
                    try:
                        if isinstance(data_arq_str, str):
                            data_arq = datetime.strptime(data_arq_str[:10], '%Y-%m-%d')
                        else:
                            data_arq = data_arq_str

                        if not (self.datas.data_inicial.date() <= data_arq.date() <= self.datas.data_final.date()):
                            continue
                    except Exception:
                        pass

                # Obtem GUID do arquivo
                arquivo_guid = arq.get('guid') or arq.get('id')
                if not arquivo_guid:
                    continue

                # Nome do arquivo
                nome_arquivo = arq.get('nome') or arq.get('nomeArquivo') or f'{uuid}_{tipo}'
                dest_file = temp_fundo / nome_arquivo

                # Baixa arquivo usando o endpoint correto
                if self.api.download_file_by_guid(uuid, arquivo_guid, dest_file):
                    # Move para destino final
                    data_ref = data_arq if data_arq else self.datas.data_inicial
                    if self.file_handler.processar_arquivo_individual(
                        dest_file, nome, data_ref, tipo, pasta
                    ):
                        qtd_baixados += 1

            if qtd_baixados > 0:
                resultado['status'] = 'sucesso'
                resultado['arquivos'] = qtd_baixados
                log.info(f"  {nome} [{tipo}]: {qtd_baixados} arquivo(s)")
            else:
                log.warning(f"  {nome} [{tipo}]: Nenhum arquivo no periodo")

            # Limpa pasta temporaria
            if temp_fundo.exists():
                shutil.rmtree(temp_fundo, ignore_errors=True)

        except Exception as e:
            log.error(f"  {nome} [{tipo}]: Erro - {e}")
            resultado['motivo'] = str(e)

        return resultado


# =============================================================================
# CLASSE PRINCIPAL: QoreAutomationAPI
# =============================================================================

class QoreAutomationAPI:
    """Classe principal de automacao do portal QORE via API."""

    def __init__(self, paths: QorePaths, credentials: QoreCredentials,
                 flags: QoreFlags, datas: QoreDatas):
        self.paths = paths
        self.credentials = credentials
        self.flags = flags
        self.datas = datas

        # Componentes
        self.api_client = QoreAPIClient(credentials)
        self.file_handler = FileHandler(paths.temp_download, paths)
        self.fundo_manager = FundoManager(paths.bd_path)
        self.download_manager = None

        # Estatisticas
        self.stats = {'total': 0, 'sucesso': 0, 'erro': 0}

    def executar(self):
        """Executa a rotina completa de automacao via API."""
        self._print_header()

        # Limpa temp
        self.file_handler.limpar_temp()

        # Carrega fundos do BD.xlsx
        if not self.fundo_manager.carregar_fundos():
            log.critical("Nenhum fundo encontrado. Abortando.")
            return

        # Autentica na API
        if not self.api_client.authenticate():
            log.critical("Falha na autenticacao. Abortando.")
            return

        # Inicializa download manager
        self.download_manager = QoreDownloadManager(
            self.api_client, self.file_handler,
            self.fundo_manager, self.datas
        )

        # Carrega UUIDs dos fundos
        if not self.download_manager.load_fundos_uuid():
            log.warning("Nao foi possivel carregar UUIDs. Tentando descobrir via API...")
            # TODO: Implementar descoberta de UUIDs via API se necessario

        # Determina tipos a baixar
        tipos_ativos = []
        if self.flags.pdf_enabled:
            tipos_ativos.append('PDF')
        if self.flags.excel_enabled:
            tipos_ativos.append('EXCEL')
        if self.flags.xml_enabled:
            tipos_ativos.append('XML')

        if not tipos_ativos:
            log.warning("Nenhum tipo de download habilitado")
            return

        log.info(f"Tipos habilitados: {', '.join(tipos_ativos)}")

        # Processa downloads
        self.stats['total'] = len(self.fundo_manager.fundos) * len(tipos_ativos)

        print()
        log.info("=" * 50)
        log.info(f"Iniciando downloads ({self.stats['total']} tarefas)...")
        log.info("=" * 50)

        resultados = self.download_manager.download_all_funds(tipos_ativos)

        # Contabiliza resultados
        for key, result in resultados.items():
            if result['status'] == 'sucesso':
                self.stats['sucesso'] += 1
            else:
                self.stats['erro'] += 1

        self._print_resumo()

    def _print_header(self):
        """Imprime cabecalho da execucao."""
        print()
        print("=" * 70)
        print("  AUTOMACAO QORE API V1 - HTTP REQUESTS")
        print("=" * 70)
        print()

        log.info(f"Data referencia: {self.datas.data_exibicao}")
        if self.datas.is_lote:
            log.info(f"Periodo: {self.datas.data_inicial.strftime('%d/%m/%Y')} a {self.datas.data_final.strftime('%d/%m/%Y')}")
        log.info(f"Workers: {APIConfig().NUM_WORKERS} threads")
        log.info(f"Tipos: PDF={self.flags.pdf_enabled} | "
                f"Excel={self.flags.excel_enabled} | "
                f"XML={self.flags.xml_enabled}")
        print()

    def _print_resumo(self):
        """Imprime resumo da execucao."""
        print()
        print("-" * 70)
        print("  RESUMO")
        print("-" * 70)
        print(f"    Total de tarefas:    {self.stats['total']}")
        print(f"    Processadas:         {self.stats['sucesso']}")
        print(f"    Com erro:            {self.stats['erro']}")
        print("=" * 70)
        print()


# =============================================================================
# FUNCAO DE CARREGAMENTO DE CONFIGURACAO
# =============================================================================

def carregar_config_planilha(caminho: str) -> Tuple[QorePaths, QoreCredentials, QoreFlags, QoreDatas]:
    """Carrega configuracoes da planilha DOWNLOADS_AUX.xlsx."""

    log.info(f"Lendo configuracoes de: {caminho}")

    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb['Downloads']

    # Datas
    data_ini = ws['C4'].value
    data_fim = ws['C5'].value

    if isinstance(data_ini, str):
        data_ini = datetime.strptime(data_ini, '%d/%m/%Y')
    if isinstance(data_fim, str):
        data_fim = datetime.strptime(data_fim, '%d/%m/%Y')
    if data_fim is None:
        data_fim = data_ini

    # Paths
    paths = QorePaths(
        pdf=str(ws['I9'].value or '').strip(),
        excel=str(ws['I13'].value or '').strip(),
        xml=str(ws['I21'].value or r'C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\03. Arquivos Rotina\XML_QORE').strip(),
        pdf_monitoramento=r'C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\03. Arquivos Rotina\05. PDF',
        temp_download=str(ws['I20'].value or '').strip(),
        bd_path=str(ws['I19'].value or '').strip()
    )

    # Credenciais
    credentials = QoreCredentials(
        url=str(ws['M10'].value or '').strip(),
        email=str(ws['N10'].value or '').strip(),
        senha=str(ws['O10'].value or '').strip()
    )

    # Flags
    flags = QoreFlags(
        qore_enabled=validar_boolean(ws['C24'].value),
        pdf_enabled=validar_boolean(ws['C25'].value),
        pdf_lote=validar_boolean(ws['C26'].value),
        excel_enabled=validar_boolean(ws['C27'].value),
        excel_lote=validar_boolean(ws['C28'].value),
        xml_enabled=validar_boolean(ws['C29'].value),
        xml_lote=validar_boolean(ws['C30'].value)
    )

    # Datas
    datas = QoreDatas(data_inicial=data_ini, data_final=data_fim)

    return paths, credentials, flags, datas


# =============================================================================
# MAIN
# =============================================================================

if __name__ == '__main__':
    DEFAULT_AUX_PATH = r'C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\03. Arquivos Rotina\07. DEPARA\DOWNLOADS_AUX.xlsx'

    # Caminho da planilha
    aux_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_AUX_PATH

    if not os.path.exists(aux_path):
        log.critical(f"Planilha nao encontrada: {aux_path}")
        sys.exit(1)

    try:
        # Carrega configuracoes
        paths, credentials, flags, datas = carregar_config_planilha(aux_path)

        # Verifica se esta habilitado
        if not flags.qore_enabled:
            log.info("Automacao QORE desabilitada na planilha (C24)")
            sys.exit(0)

        # Verifica se tem algo para baixar
        if not any([flags.pdf_enabled, flags.excel_enabled, flags.xml_enabled]):
            log.info("Nenhum tipo de download habilitado")
            sys.exit(0)

        # Cria pastas necessarias
        os.makedirs(paths.temp_download, exist_ok=True)
        os.makedirs(paths.pdf_monitoramento, exist_ok=True)
        if paths.xml:
            os.makedirs(paths.xml, exist_ok=True)

        # Executa
        bot = QoreAutomationAPI(paths, credentials, flags, datas)
        bot.executar()

    except Exception as e:
        log.critical(f"Erro fatal: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
