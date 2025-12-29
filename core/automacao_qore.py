"""
================================================================================
AUTOMACAO QORE V14 - MULTIDRIVER PARALELO
================================================================================

Melhorias V14 sobre V13:
- Downloads VERDADEIRAMENTE paralelos com MULTIPLOS DRIVERS Chrome
- 10 instancias Chrome simultaneas (ThreadPoolExecutor)
- Cada worker processa fundos independentemente
- Speedup ~10x em relacao a versao sequencial

Melhorias V13 sobre V12:
- Downloads em PARALELO usando multiplas abas do navegador

Melhorias V12 sobre V11:
- Chrome otimizado (page_load_strategy='eager', imagens desativadas)
- Retry automatico com backoff exponencial (2 tentativas)
- Session recovery automatico

Autor: ETL Team
Data: Dezembro 2025
================================================================================
"""

import os
import re
import sys
import time
import shutil
import logging
import zipfile
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

import pandas as pd
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException
)


# =============================================================================
# CONFIGURACAO DE LOGGING
# =============================================================================

class LogFormatter(logging.Formatter):
    """Formatter customizado com cores e simbolos."""

    COLORS = {
        'DEBUG': '\033[36m',     # Cyan
        'INFO': '\033[32m',      # Green
        'WARNING': '\033[33m',   # Yellow
        'ERROR': '\033[31m',     # Red
        'CRITICAL': '\033[35m',  # Magenta
        'RESET': '\033[0m'
    }

    SYMBOLS = {
        'DEBUG': '[.]',
        'INFO': '[+]',
        'WARNING': '[!]',
        'ERROR': '[X]',
        'CRITICAL': '[!!!]'
    }

    def format(self, record):
        symbol = self.SYMBOLS.get(record.levelname, '[?]')
        color = self.COLORS.get(record.levelname, '')
        reset = self.COLORS['RESET']

        # Formato: [SYMBOL] MENSAGEM
        record.msg = f"{color}{symbol}{reset} {record.msg}"
        return super().format(record)


def setup_logging(level=logging.INFO) -> logging.Logger:
    """Configura e retorna o logger principal."""
    logger = logging.getLogger('QoreAutomation')
    logger.setLevel(level)

    if not logger.handlers:
        handler = logging.StreamHandler(sys.stdout)
        handler.setFormatter(LogFormatter('%(message)s'))
        logger.addHandler(handler)

    return logger


log = setup_logging()


# =============================================================================
# CONSTANTES E CONFIGURACOES (Centralizadas)
# =============================================================================

@dataclass
class Timeouts:
    """Timeouts centralizados (em segundos) - OTIMIZADOS."""
    PAGE_LOAD: int = 8
    ELEMENT_CLICK: int = 6
    ELEMENT_PRESENT: int = 4
    DOWNLOAD_WAIT: int = 45
    DOWNLOAD_CHECK_INTERVAL: float = 0.3
    POST_CLICK_WAIT: int = 1
    POST_NAVIGATION_WAIT: int = 1
    MAX_RETRIES: int = 2  # Tentativas em caso de falha
    NUM_WORKERS: int = 10  # V14: 10 Chromes simultaneos


@dataclass
class ReportConfig:
    """Configuracao para cada tipo de relatorio."""
    extension: str
    type_name: str
    button_text: str

    def get_filename(self, fundo: str, data: datetime) -> str:
        """Gera o nome base do arquivo."""
        return f"{data.strftime('%d.%m')} - {self.type_name} - {fundo}"


# Configuracoes dos tipos de relatorio
REPORT_CONFIGS: Dict[str, ReportConfig] = {
    'PDF': ReportConfig('.pdf', 'Carteira Diaria', 'Carteira PDF'),
    'EXCEL': ReportConfig('.xlsx', 'Carteira Excel', 'Carteira Excel'),
    'XML': ReportConfig('.xml', 'Carteira XML', 'XML Anbima 5.0')
}

# Mapeamento de fundos BLOKO para padroes de busca em arquivos
BLOKO_PATTERNS: Dict[str, str] = {
    'BLOKO URBANISMO': 'urbanismo',
    'BLOKO FIM': 'fundo-de-investimento'
}

# Meses por extenso
MESES_EXTENSO: Dict[str, str] = {
    '01': 'Janeiro', '02': 'Fevereiro', '03': 'Marco', '04': 'Abril',
    '05': 'Maio', '06': 'Junho', '07': 'Julho', '08': 'Agosto',
    '09': 'Setembro', '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro'
}


# =============================================================================
# DATACLASSES DE CONFIGURACAO
# =============================================================================

@dataclass
class QoreCredentials:
    """Credenciais de acesso ao QORE."""
    url: str
    email: str
    senha: str


@dataclass
class QorePaths:
    """Caminhos de destino dos arquivos."""
    pdf: str = ''
    excel: str = ''
    xml: str = ''
    pdf_monitoramento: str = ''
    temp_download: str = ''
    bd_path: str = ''
    base_fundos: str = r'C:\bloko\Fundos - Documentos'


@dataclass
class QoreFlags:
    """Flags de controle da automacao."""
    qore_enabled: bool = False
    pdf_enabled: bool = False
    pdf_lote: bool = False
    excel_enabled: bool = False
    excel_lote: bool = False
    xml_enabled: bool = False
    xml_lote: bool = False


@dataclass
class QoreDatas:
    """Datas de referencia."""
    data_inicial: datetime = None
    data_final: datetime = None

    @property
    def data_exibicao(self) -> str:
        """Data formatada para exibicao (dd/mm/yyyy)."""
        return self.data_inicial.strftime('%d/%m/%Y') if self.data_inicial else ''

    @property
    def is_lote(self) -> bool:
        """Verifica se e modo lote (datas diferentes)."""
        return self.data_inicial != self.data_final


# =============================================================================
# FUNCOES UTILITARIAS
# =============================================================================

def validar_boolean(valor) -> bool:
    """Converte valores da planilha para booleano."""
    if pd.isna(valor) or valor is None:
        return False
    valor_str = str(valor).strip().upper()
    return valor_str in {'SIM', 'S', 'TRUE', 'VERDADEIRO', 'YES', 'Y', '1'}


def extrair_data_de_nome_arquivo(nome_arquivo: str) -> Optional[datetime]:
    """
    Extrai data do nome do arquivo (formato YYYYMMDD).
    Usa regex com underscore para evitar confundir com CNPJ.
    """
    # Primeiro tenta com underscore (mais seguro)
    matches = re.findall(r'_(\d{8})', nome_arquivo)

    # Fallback: qualquer sequencia de 8 digitos
    if not matches:
        matches = re.findall(r'(\d{8})', nome_arquivo)

    for date_str in matches:
        try:
            dt = datetime.strptime(date_str, '%Y%m%d')
            # Valida ano razoavel (evita CNPJ)
            if 2000 <= dt.year <= 2035:
                return dt
        except ValueError:
            continue

    return None


def get_versioned_filepath(directory: str, base_name: str, extension: str) -> str:
    """
    Gera caminho com versionamento automatico.
    Ex: arquivo.pdf -> arquivo (1).pdf -> arquivo (2).pdf
    """
    target = os.path.join(directory, f"{base_name}{extension}")

    if not os.path.exists(target):
        return target

    version = 1
    while True:
        versioned = os.path.join(directory, f"{base_name} ({version}){extension}")
        if not os.path.exists(versioned):
            return versioned
        version += 1


# =============================================================================
# CLASSE: FundoManager
# =============================================================================

class FundoManager:
    """Gerencia a lista de fundos e suas siglas."""

    def __init__(self, bd_path: str):
        self.bd_path = bd_path
        self.fundos: Dict[str, str] = {}  # {nome_fundo: pasta_destino}
        self.siglas: Dict[str, str] = {}  # {nome_fundo: sigla_busca}

    def carregar_fundos(self) -> bool:
        """Carrega fundos do arquivo BD.xlsx (otimizado)."""
        try:
            # Otimizado: le apenas colunas necessarias (B=1, C=2, J=9)
            df = pd.read_excel(
                self.bd_path,
                sheet_name='BD',
                engine='openpyxl',
                header=None,
                usecols=[1, 2, 9],
                dtype=str
            )

            for _, row in df.iterrows():
                flag_qore = str(row.iloc[2] or '').strip().upper()  # Coluna J (indice 2 apos usecols)

                if flag_qore in {'SIM', 'S', 'TRUE', 'YES', 'VERDADEIRO', 'QORE'}:
                    apelido = row.iloc[0]  # Coluna B (indice 0 apos usecols)
                    caminho = row.iloc[1]  # Coluna C (indice 1 apos usecols)

                    if pd.notna(apelido) and pd.notna(caminho):
                        apelido_clean = str(apelido).strip()
                        caminho_clean = str(caminho).strip()

                        if apelido_clean and caminho_clean:
                            # Tratamento especial BLOKO
                            nome_final = self._processar_nome_bloko(apelido_clean)
                            self.fundos[nome_final] = caminho_clean

            # Gera siglas
            self._gerar_siglas()

            log.info(f"Carregados {len(self.fundos)} fundos QORE")
            return len(self.fundos) > 0

        except Exception as e:
            log.error(f"Falha ao ler BD.xlsx: {e}")
            return False

    def _processar_nome_bloko(self, apelido: str) -> str:
        """Processa nomes especiais de fundos BLOKO."""
        partes = apelido.split()

        if len(partes) > 1 and partes[1] == 'BLOKO':
            if partes[0] == 'FIP':
                return 'BLOKO URBANISMO'
            else:
                return 'BLOKO FIM'

        return apelido

    def _gerar_siglas(self):
        """Gera mapa de siglas para busca."""
        for nome in self.fundos:
            if 'BLOKO' in nome.upper():
                # BLOKO usa nome completo
                self.siglas[nome] = nome
            else:
                # Outros: extrai segunda palavra (ex: "FIDC ABC" -> "ABC")
                partes = nome.split(' ', 1)
                self.siglas[nome] = partes[1].strip() if len(partes) > 1 else nome

    def get_sigla(self, nome_fundo: str) -> str:
        """Retorna a sigla de busca para um fundo."""
        return self.siglas.get(nome_fundo, nome_fundo)

    def is_bloko(self, nome_fundo: str) -> bool:
        """Verifica se e um fundo BLOKO."""
        return 'BLOKO' in nome_fundo.upper()

    def get_bloko_pattern(self, nome_fundo: str) -> str:
        """Retorna o padrao de busca em arquivos para fundos BLOKO."""
        return BLOKO_PATTERNS.get(nome_fundo, '')


# =============================================================================
# CLASSE: FileHandler (Unificado - Corrige ponto 2)
# =============================================================================

class FileHandler:
    """
    Gerencia download e movimentacao de arquivos.
    UNIFICADO: mesma logica para fundos normais e BLOKO.
    """

    def __init__(self, temp_path: str, paths: QorePaths):
        self.temp_path = temp_path
        self.paths = paths
        self.timeouts = Timeouts()

    def limpar_temp(self):
        """Limpa pasta temporaria de downloads."""
        if not os.path.exists(self.temp_path):
            os.makedirs(self.temp_path, exist_ok=True)
            return

        for item in os.listdir(self.temp_path):
            item_path = os.path.join(self.temp_path, item)
            try:
                if os.path.isfile(item_path):
                    os.unlink(item_path)
                elif os.path.isdir(item_path):
                    shutil.rmtree(item_path)
            except Exception:
                pass

        log.debug(f"Pasta temp limpa: {self.temp_path}")

    def aguardar_download(self, extension: str, sigla: str = '', timeout: int = None) -> Optional[Path]:
        """
        Aguarda arquivo aparecer na pasta temp.
        Usa polling inteligente em vez de sleep fixo.
        """
        timeout = timeout or self.timeouts.DOWNLOAD_WAIT
        end_time = datetime.now().timestamp() + timeout

        while datetime.now().timestamp() < end_time:
            # Lista arquivos validos (nao temporarios)
            arquivos = [
                f for f in Path(self.temp_path).glob(f'*{extension}')
                if not f.name.endswith(('.crdownload', '.tmp', '.partial'))
            ]

            if arquivos:
                # Ordena por data de modificacao (mais recente primeiro)
                arquivos.sort(key=lambda x: x.stat().st_mtime, reverse=True)

                # Se tem sigla, filtra
                if sigla:
                    for arq in arquivos:
                        if sigla.lower() in arq.name.lower():
                            # Verifica se download completou (tamanho estavel)
                            if self._download_completo(arq):
                                return arq
                else:
                    if self._download_completo(arquivos[0]):
                        return arquivos[0]

            # Polling interval
            time.sleep(self.timeouts.DOWNLOAD_CHECK_INTERVAL)

        return None

    def _download_completo(self, arquivo: Path) -> bool:
        """
        Verifica se o download do arquivo completou.
        Otimizado: verifica .crdownload primeiro (sem sleep).
        """
        try:
            # Verifica se existe arquivo temporario do Chrome (.crdownload)
            crdownload = Path(str(arquivo) + '.crdownload')
            if crdownload.exists():
                return False

            # Verifica tamanho minimo
            size1 = arquivo.stat().st_size
            if size1 == 0:
                return False

            # Verifica estabilidade do tamanho (sleep reduzido)
            time.sleep(0.2)
            size2 = arquivo.stat().st_size
            return size1 == size2

        except Exception:
            return False

    def processar_zip_lote(self, fundo_nome: str, fundo_manager: FundoManager,
                          report_type: str, data_referencia: datetime) -> int:
        """
        Processa arquivos de download em lote (ZIP).
        UNIFICADO: mesma logica para BLOKO e outros fundos.
        """
        # Encontra o ZIP
        zips = list(Path(self.temp_path).glob('*.zip'))
        if not zips:
            log.warning("Nenhum ZIP encontrado na pasta temp")
            return 0

        zip_file = max(zips, key=lambda x: x.stat().st_mtime)
        log.info(f"Extraindo ZIP: {zip_file.name}")

        # Cria pasta temp de extracao
        temp_extract = Path(self.temp_path) / 'temp_extract'
        temp_extract.mkdir(exist_ok=True)

        try:
            # Extrai
            with zipfile.ZipFile(zip_file, 'r') as zf:
                zf.extractall(temp_extract)

            # Remove ZIP
            zip_file.unlink()

            # Processa arquivos extraidos
            config = REPORT_CONFIGS.get(report_type.upper())
            if not config:
                return 0

            arquivos_processados = 0

            # Define padrao de busca
            if fundo_manager.is_bloko(fundo_nome):
                padrao = fundo_manager.get_bloko_pattern(fundo_nome)
            else:
                padrao = fundo_manager.get_sigla(fundo_nome).lower()

            for arquivo in temp_extract.glob(f'*{config.extension}'):
                # Verifica se arquivo pertence ao fundo
                if padrao and padrao.lower() not in arquivo.name.lower():
                    continue

                # Extrai data do nome do arquivo
                data_arquivo = extrair_data_de_nome_arquivo(arquivo.name)
                if not data_arquivo:
                    data_arquivo = data_referencia

                # Move para destinos
                if self._mover_arquivo(
                    arquivo, fundo_nome, data_arquivo, report_type,
                    fundo_manager.fundos.get(fundo_nome, '')
                ):
                    arquivos_processados += 1

            return arquivos_processados

        finally:
            # Limpa pasta de extracao
            if temp_extract.exists():
                shutil.rmtree(temp_extract, ignore_errors=True)

    def processar_arquivo_individual(self, arquivo: Path, fundo_nome: str,
                                    data: datetime, report_type: str,
                                    pasta_fundo: str) -> bool:
        """Processa um arquivo individual (modo nao-lote)."""
        return self._mover_arquivo(arquivo, fundo_nome, data, report_type, pasta_fundo)

    def _mover_arquivo(self, arquivo: Path, fundo_nome: str, data: datetime,
                      report_type: str, pasta_fundo: str) -> bool:
        """
        Move arquivo para destino(s) final(is).
        UNIFICADO: mesma logica para todos os tipos.
        """
        config = REPORT_CONFIGS.get(report_type.upper())
        if not config:
            return False

        try:
            # Gera nome do arquivo
            nome_base = config.get_filename(fundo_nome, data)

            # Determina destinos
            destinos = self._get_destinos(report_type, data, pasta_fundo)

            primeiro_destino = None

            for i, destino_dir in enumerate(destinos):
                os.makedirs(destino_dir, exist_ok=True)

                caminho_final = get_versioned_filepath(
                    destino_dir, nome_base, config.extension
                )

                if i == 0:
                    # Primeiro destino: move
                    shutil.move(str(arquivo), caminho_final)
                    primeiro_destino = caminho_final
                    log.info(f"{report_type} movido para: {Path(caminho_final).name}")
                else:
                    # Demais destinos: copia
                    if primeiro_destino and os.path.exists(primeiro_destino):
                        shutil.copy(primeiro_destino, caminho_final)
                        log.info(f"{report_type} copiado para: {Path(caminho_final).parent.name}")

            return True

        except Exception as e:
            log.error(f"Falha ao mover arquivo: {e}")
            return False

    def _get_destinos(self, report_type: str, data: datetime, pasta_fundo: str) -> List[str]:
        """Retorna lista de diretorios de destino."""
        destinos = []
        ano = data.strftime('%Y')
        mes = data.strftime('%m')

        if report_type.upper() == 'PDF':
            # Destino 1: Pasta do fundo
            if pasta_fundo:
                mes_formatado = f"{mes} - {MESES_EXTENSO.get(mes, '')}"
                path_fundo = os.path.join(
                    self.paths.base_fundos, pasta_fundo,
                    '06. Carteiras', ano, mes_formatado
                )
                destinos.append(path_fundo)

            # Destino 2: Monitoramento
            if self.paths.pdf_monitoramento:
                destinos.append(self.paths.pdf_monitoramento)

        elif report_type.upper() == 'EXCEL':
            if self.paths.excel:
                destinos.append(self.paths.excel)

        elif report_type.upper() == 'XML':
            if self.paths.xml:
                destinos.append(self.paths.xml)

        return destinos


# =============================================================================
# CLASSE: SeleniumDriver (Waits inteligentes - Corrige ponto 7)
# =============================================================================

class SeleniumDriver:
    """
    Wrapper para Selenium com waits inteligentes.
    Evita time.sleep() fixos usando WebDriverWait.
    """

    def __init__(self, download_path: str):
        self.download_path = download_path
        self.driver = None
        self.timeouts = Timeouts()

    def iniciar(self):
        """Inicia o Chrome Driver com configuracoes otimizadas."""
        log.info("Inicializando Chrome Driver...")

        chrome_options = Options()

        # Preferencias de download
        prefs = {
            'download.default_directory': self.download_path,
            'download.prompt_for_download': False,
            'plugins.always_open_pdf_externally': True,
            'download.directory_upgrade': True,
            'safebrowsing.enabled': True,
            'profile.default_content_settings.popups': 0
        }
        chrome_options.add_experimental_option('prefs', prefs)

        # Argumentos de estabilidade
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--start-maximized')

        # Argumentos de performance
        chrome_options.add_argument('--disable-extensions')
        chrome_options.add_argument('--disable-infobars')
        chrome_options.add_argument('--disable-notifications')
        chrome_options.add_argument('--disable-logging')
        chrome_options.add_argument('--log-level=3')
        chrome_options.add_argument('--blink-settings=imagesEnabled=false')

        # Estrategia de carregamento mais rapida
        chrome_options.page_load_strategy = 'eager'

        try:
            self.driver = webdriver.Chrome(options=chrome_options)
            self.driver.implicitly_wait(self.timeouts.ELEMENT_PRESENT)
            log.info("Chrome Driver iniciado com sucesso")
        except Exception as e:
            log.critical(f"Falha ao iniciar Chrome: {e}")
            raise

    def fechar(self):
        """Fecha o driver."""
        if self.driver:
            try:
                self.driver.quit()
                log.info("Chrome Driver encerrado")
            except Exception:
                pass
            finally:
                self.driver = None

    def navegar(self, url: str) -> bool:
        """Navega para uma URL e aguarda carregamento."""
        try:
            self.driver.get(url)
            self._wait_page_load()
            return True
        except Exception as e:
            log.error(f"Falha ao navegar para {url}: {e}")
            return False

    def _wait_page_load(self):
        """Aguarda pagina carregar completamente."""
        WebDriverWait(self.driver, self.timeouts.PAGE_LOAD).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )

    def clicar_elemento(self, by: By, value: str, timeout: int = None) -> bool:
        """Clica em elemento com wait inteligente."""
        timeout = timeout or self.timeouts.ELEMENT_CLICK

        try:
            elemento = WebDriverWait(self.driver, timeout).until(
                EC.element_to_be_clickable((by, value))
            )

            # Scroll para elemento se necessario
            self.driver.execute_script(
                "arguments[0].scrollIntoView({block: 'center'});", elemento
            )

            elemento.click()
            return True

        except TimeoutException:
            log.warning(f"Elemento nao encontrado: {value}")
            return False
        except StaleElementReferenceException:
            # Retry uma vez
            return self._retry_click(by, value, timeout)
        except Exception as e:
            log.error(f"Erro ao clicar: {e}")
            return False

    def _retry_click(self, by: By, value: str, timeout: int) -> bool:
        """Retry de clique em caso de stale element."""
        try:
            elemento = WebDriverWait(self.driver, timeout).until(
                EC.element_to_be_clickable((by, value))
            )
            elemento.click()
            return True
        except Exception:
            return False

    def encontrar_elemento(self, by: By, value: str, timeout: int = None):
        """Encontra elemento com wait."""
        timeout = timeout or self.timeouts.ELEMENT_PRESENT

        try:
            return WebDriverWait(self.driver, timeout).until(
                EC.presence_of_element_located((by, value))
            )
        except TimeoutException:
            return None

    def encontrar_elementos(self, by: By, value: str) -> list:
        """Encontra multiplos elementos."""
        try:
            return self.driver.find_elements(by, value)
        except Exception:
            return []

    def preencher_campo(self, by: By, value: str, texto: str) -> bool:
        """Preenche campo de texto."""
        elemento = self.encontrar_elemento(by, value)
        if elemento:
            try:
                elemento.clear()
                elemento.send_keys(texto)
                return True
            except Exception:
                pass
        return False

    def preencher_data_js(self, element_id: str, data: datetime) -> bool:
        """Preenche campo de data usando JavaScript (mais confiavel)."""
        try:
            data_iso = data.strftime('%Y-%m-%d')
            script = f'''
                var campo = document.getElementById('{element_id}');
                if (campo) {{
                    campo.value = '{data_iso}';
                    ['focus', 'input', 'change', 'blur'].forEach(function(evt) {{
                        campo.dispatchEvent(new Event(evt, {{bubbles: true}}));
                    }});
                    return true;
                }}
                return false;
            '''
            return self.driver.execute_script(script)
        except Exception as e:
            log.error(f"Erro ao preencher data: {e}")
            return False

    def aguardar_url_conter(self, texto: str, timeout: int = None) -> bool:
        """Aguarda URL conter determinado texto."""
        timeout = timeout or self.timeouts.PAGE_LOAD

        try:
            WebDriverWait(self.driver, timeout).until(EC.url_contains(texto))
            return True
        except TimeoutException:
            return False

    def screenshot(self, nome: str):
        """Salva screenshot para debug."""
        try:
            path = os.path.join(self.download_path, f"{nome}.png")
            self.driver.save_screenshot(path)
            log.debug(f"Screenshot salvo: {path}")
        except Exception:
            pass

    def aguardar_carregamento(self):
        """
        Aguarda carregamento da pagina usando waits inteligentes.
        Substitui time.sleep() fixos por espera real de elementos.
        """
        try:
            # Aguarda document ready
            self._wait_page_load()

            # Aguarda spinners/loaders desaparecerem (timeout curto)
            try:
                WebDriverWait(self.driver, 1).until(
                    EC.invisibility_of_element_located((By.CLASS_NAME, 'spinner'))
                )
            except TimeoutException:
                pass  # Nao tem spinner, ok

        except Exception:
            # Apenas em caso de erro, pequeno delay de fallback
            time.sleep(0.5)

    def aguardar_elemento_visivel(self, by: By, value: str, timeout: int = None) -> bool:
        """Aguarda elemento ficar visivel na pagina."""
        timeout = timeout or self.timeouts.ELEMENT_PRESENT

        try:
            WebDriverWait(self.driver, timeout).until(
                EC.visibility_of_element_located((by, value))
            )
            return True
        except TimeoutException:
            return False

    def aguardar_breve(self):
        """
        Aguarda breve para operacoes que precisam de pequeno delay.
        Usado apos preenchimento de campos, etc.
        """
        time.sleep(0.5)

    # =========================================================================
    # METODOS PARA GERENCIAMENTO DE ABAS (V13 - Paralelo)
    # =========================================================================

    def abrir_nova_aba(self) -> str:
        """Abre uma nova aba e retorna o handle."""
        self.driver.execute_script("window.open('');")
        nova_aba = self.driver.window_handles[-1]
        return nova_aba

    def trocar_para_aba(self, handle: str):
        """Troca para uma aba especifica pelo handle."""
        self.driver.switch_to.window(handle)

    def fechar_aba_atual(self):
        """Fecha a aba atual."""
        self.driver.close()

    def get_aba_principal(self) -> str:
        """Retorna o handle da aba principal (primeira)."""
        return self.driver.window_handles[0]

    def get_todas_abas(self) -> list:
        """Retorna lista de handles de todas as abas."""
        return self.driver.window_handles

    def fechar_todas_abas_extras(self):
        """Fecha todas as abas exceto a principal."""
        principal = self.get_aba_principal()
        for handle in self.driver.window_handles:
            if handle != principal:
                self.driver.switch_to.window(handle)
                self.driver.close()
        self.driver.switch_to.window(principal)

    def get_cookies(self) -> list:
        """Retorna cookies da sessao atual."""
        return self.driver.get_cookies()


# =============================================================================
# CLASSE: WorkerChrome (V14 - Multidriver Paralelo)
# =============================================================================

class WorkerChrome:
    """
    Worker que executa em thread separada com seu proprio Chrome.
    Usado para downloads verdadeiramente paralelos.
    Cada worker faz seu proprio login (mais robusto que cookies).
    Cada worker tem sua propria pasta de download (evita conflitos).
    """

    def __init__(self, worker_id: int, base_temp_path: str, credentials: QoreCredentials,
                 datas, report_config: ReportConfig):
        self.worker_id = worker_id
        self.base_temp_path = base_temp_path
        # Cria pasta exclusiva para este worker
        self.temp_path = str(Path(base_temp_path) / f'worker_{worker_id}')
        Path(self.temp_path).mkdir(parents=True, exist_ok=True)
        self.credentials = credentials
        self.datas = datas
        self.report_config = report_config
        self.driver = None
        self.timeouts = Timeouts()
        self.lock = threading.Lock()

    def iniciar(self) -> bool:
        """Inicia Chrome e faz login proprio."""
        try:
            chrome_options = Options()

            # Preferencias de download
            prefs = {
                'download.default_directory': self.temp_path,
                'download.prompt_for_download': False,
                'plugins.always_open_pdf_externally': True,
                'download.directory_upgrade': True,
                'safebrowsing.enabled': True,
                'profile.default_content_settings.popups': 0
            }
            chrome_options.add_experimental_option('prefs', prefs)

            # Argumentos de performance (mais leve para workers)
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--window-size=1200,800')
            chrome_options.add_argument('--disable-extensions')
            chrome_options.add_argument('--disable-infobars')
            chrome_options.add_argument('--disable-notifications')
            chrome_options.add_argument('--disable-logging')
            chrome_options.add_argument('--log-level=3')
            chrome_options.add_argument('--blink-settings=imagesEnabled=false')
            chrome_options.page_load_strategy = 'eager'

            self.driver = webdriver.Chrome(options=chrome_options)
            self.driver.implicitly_wait(self.timeouts.ELEMENT_PRESENT)

            # Faz login proprio
            if self._fazer_login():
                log.info(f"  Worker {self.worker_id}: Chrome iniciado e logado")
                return True
            else:
                log.error(f"  Worker {self.worker_id}: Falha no login")
                return False

        except Exception as e:
            log.error(f"  Worker {self.worker_id}: Falha ao iniciar - {e}")
            return False

    def _fazer_login(self) -> bool:
        """Faz login no QORE."""
        try:
            self.driver.get(self.credentials.url)
            time.sleep(1)

            # Email
            email_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.NAME, 'email'))
            )
            email_field.clear()
            email_field.send_keys(self.credentials.email)

            # Senha + Enter
            senha_field = self.driver.find_element(By.NAME, 'password')
            senha_field.clear()
            senha_field.send_keys(self.credentials.senha + Keys.RETURN)

            # Aguarda dashboard carregar
            WebDriverWait(self.driver, 15).until(
                lambda d: 'dashboard' in d.current_url.lower()
            )

            return True

        except Exception as e:
            log.error(f"  Worker {self.worker_id}: Erro login - {e}")
            return False

    def processar_fundo(self, url: str, sigla: str, nome_fundo: str) -> dict:
        """Processa um fundo: navega, clica, baixa."""
        resultado = {
            'nome': nome_fundo,
            'sigla': sigla,
            'status': 'erro',
            'worker': self.worker_id
        }

        try:
            # Navega para URL do fundo
            self.driver.get(url)
            time.sleep(0.5)

            # Clica no botao do tipo de relatorio
            botao_xpath = f"//button[contains(., '{self.report_config.button_text}')]"
            try:
                botao = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, botao_xpath))
                )
                botao.click()
                time.sleep(0.5)
            except Exception:
                log.warning(f"  Worker {self.worker_id}: Botao nao encontrado para {sigla}")
                return resultado

            # Inicia download em lote
            if self._iniciar_download_lote():
                # Aguarda download completar
                if self._aguardar_download():
                    resultado['status'] = 'sucesso'
                    log.info(f"  Worker {self.worker_id}: {sigla} - OK!")
                else:
                    log.warning(f"  Worker {self.worker_id}: {sigla} - Timeout download")
            else:
                log.warning(f"  Worker {self.worker_id}: {sigla} - Falha ao iniciar")

        except Exception as e:
            log.error(f"  Worker {self.worker_id}: Erro {sigla} - {e}")

        return resultado

    def _iniciar_download_lote(self) -> bool:
        """Inicia download em lote com WebDriverWait para garantir elementos prontos."""
        try:
            # 1. Clica no menu (...) - aguarda estar clicavel
            try:
                menus = self.driver.find_elements(By.XPATH, '//div[@data-kt-menu-trigger="click"]')
                for menu in menus:
                    try:
                        icone = menu.find_element(By.TAG_NAME, 'i')
                        if 'ellipsis' in icone.get_attribute('class'):
                            menu.click()
                            break
                    except Exception:
                        continue
            except Exception as e:
                log.warning(f"    Worker {self.worker_id}: Menu nao encontrado - {e}")
                return False

            time.sleep(0.5)

            # 2. Aguarda e clica em "Download em Lote"
            try:
                link_download = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, '//a[contains(., "Download em Lote")]'))
                )
                link_download.click()
            except Exception as e:
                log.warning(f"    Worker {self.worker_id}: Link Download em Lote nao encontrado - {e}")
                return False

            time.sleep(0.5)

            # 3. Aguarda modal e preenche datas
            try:
                WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.ID, 'dataInicial'))
                )
            except Exception as e:
                log.warning(f"    Worker {self.worker_id}: Campo dataInicial nao encontrado - {e}")
                return False

            # Preenche datas via JS com eventos (para frameworks React/Vue)
            data_ini = self.datas.data_inicial.strftime('%Y-%m-%d')
            data_fim = self.datas.data_final.strftime('%Y-%m-%d')

            self.driver.execute_script(f"""
                function setValueWithEvents(element, value) {{
                    var nativeInputValueSetter = Object.getOwnPropertyDescriptor(
                        window.HTMLInputElement.prototype, 'value'
                    ).set;
                    nativeInputValueSetter.call(element, value);
                    element.dispatchEvent(new Event('input', {{ bubbles: true }}));
                    element.dispatchEvent(new Event('change', {{ bubbles: true }}));
                }}
                var dataIni = document.getElementById('dataInicial');
                var dataFim = document.getElementById('dataFinal');
                if (dataIni) {{ setValueWithEvents(dataIni, '{data_ini}'); }}
                if (dataFim) {{ setValueWithEvents(dataFim, '{data_fim}'); }}
            """)

            time.sleep(0.5)

            # 4. Clica no botao Download do modal
            try:
                # Tenta xpath mais especifico primeiro (botao primario do modal)
                botao_download = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH,
                        '//div[contains(@class, "modal")]//button[contains(., "Download")]'))
                )
                botao_download.click()
                log.info(f"    Worker {self.worker_id}: Download em lote solicitado")
            except Exception:
                # Fallback: tenta qualquer botao Download
                try:
                    botao = self.driver.find_element(By.XPATH, '//button[contains(., "Download")]')
                    self.driver.execute_script("arguments[0].click();", botao)
                    log.info(f"    Worker {self.worker_id}: Download via JS")
                except Exception as e2:
                    log.warning(f"    Worker {self.worker_id}: Botao Download nao encontrado - {e2}")
                    return False

            return True

        except Exception as e:
            log.error(f"  Worker {self.worker_id}: Erro download lote - {e}")
            return False

    def _aguardar_download(self, timeout: int = 60) -> bool:
        """Aguarda download do ZIP completar."""
        start = time.time()
        download_detectado = False

        while time.time() - start < timeout:
            # Verifica se tem .crdownload (download em progresso)
            crdownloads = list(Path(self.temp_path).glob('*.crdownload'))
            zips = list(Path(self.temp_path).glob('*.zip'))

            # Log de progresso (apenas primeira vez que detecta download)
            if crdownloads and not download_detectado:
                download_detectado = True
                log.info(f"    Worker {self.worker_id}: Download iniciado...")

            # Se tem ZIP e nao tem crdownload, download completou
            if zips and not crdownloads:
                return True

            time.sleep(0.5)

        # Log de debug se timeout
        crdownloads = list(Path(self.temp_path).glob('*.crdownload'))
        zips = list(Path(self.temp_path).glob('*.zip'))
        log.warning(f"    Worker {self.worker_id}: Timeout - ZIPs={len(zips)}, .crdownload={len(crdownloads)}")

        return False

    def fechar(self):
        """Fecha o Chrome deste worker."""
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass


# =============================================================================
# CLASSE PRINCIPAL: QoreAutomation
# =============================================================================

class QoreAutomation:
    """
    Classe principal de automacao do portal QORE.
    Orquestra login, navegacao e downloads.
    """

    def __init__(self, paths: QorePaths, credentials: QoreCredentials,
                 flags: QoreFlags, datas: QoreDatas):
        self.paths = paths
        self.credentials = credentials
        self.flags = flags
        self.datas = datas

        # Componentes
        self.selenium = SeleniumDriver(paths.temp_download)
        self.file_handler = FileHandler(paths.temp_download, paths)
        self.fundo_manager = FundoManager(paths.bd_path)

        # Estatisticas
        self.stats = {'total': 0, 'sucesso': 0, 'erro': 0}

        # Configuracoes de retry
        self.timeouts = Timeouts()

    def _executar_com_retry(self, func, descricao: str, *args, **kwargs):
        """
        Executa uma funcao com retry e backoff exponencial.
        Retorna o resultado da funcao ou None em caso de falha total.
        """
        max_retries = self.timeouts.MAX_RETRIES

        for attempt in range(max_retries + 1):
            try:
                return func(*args, **kwargs)
            except Exception as e:
                if attempt < max_retries:
                    wait_time = 2 ** attempt  # 1s, 2s, 4s...
                    log.warning(f"{descricao}: tentativa {attempt + 1} falhou, "
                               f"retry em {wait_time}s... ({e})")
                    time.sleep(wait_time)
                else:
                    log.error(f"{descricao}: falha apos {max_retries + 1} tentativas - {e}")
                    raise

    def executar(self):
        """Executa a rotina completa de automacao V14 - Multidriver."""
        self._print_header()

        # Limpa temp
        self.file_handler.limpar_temp()

        # Carrega fundos
        if not self.fundo_manager.carregar_fundos():
            log.critical("Nenhum fundo encontrado. Abortando.")
            return

        # Inicia Selenium principal (para login e coleta de URLs)
        try:
            self.selenium.iniciar()
        except Exception:
            return

        try:
            # Login
            if not self._fazer_login():
                log.critical("Falha no login. Abortando.")
                return

            # V14: Processa fundos com multiplos drivers
            self._processar_fundos_v14()

        finally:
            self.selenium.fechar()
            self._print_resumo()

    def _print_header(self):
        """Imprime cabecalho da execucao."""
        print()
        print("=" * 70)
        print("  AUTOMACAO QORE V14 - MULTIDRIVER PARALELO")
        print("=" * 70)
        print()

        log.info(f"Data referencia: {self.datas.data_exibicao}")
        log.info(f"Workers: {self.timeouts.NUM_WORKERS} Chromes simultaneos")
        log.info(f"Tipos: PDF={self.flags.pdf_enabled} | "
                f"Excel={self.flags.excel_enabled} | "
                f"XML={self.flags.xml_enabled}")
        print()

    def _print_resumo(self):
        """Imprime resumo da execucao."""
        print()
        print("-" * 70)
        print("  RESUMO")
        print("-" * 70)
        print(f"    Total de fundos:     {self.stats['total']}")
        print(f"    Processados:         {self.stats['sucesso']}")
        print(f"    Com erro:            {self.stats['erro']}")
        print("=" * 70)
        print()

    def _fazer_login(self) -> bool:
        """Realiza login no portal QORE com retry automatico."""
        max_retries = self.timeouts.MAX_RETRIES

        for attempt in range(max_retries + 1):
            try:
                log.info(f"Acessando: {self.credentials.url}")

                if not self.selenium.navegar(self.credentials.url):
                    raise Exception("Falha ao navegar para URL")

                log.info("Inserindo credenciais...")

                # Email
                if not self.selenium.preencher_campo(By.NAME, 'email', self.credentials.email):
                    raise Exception("Campo email nao encontrado")

                # Senha + Enter
                elem_senha = self.selenium.encontrar_elemento(By.NAME, 'password')
                if elem_senha:
                    elem_senha.clear()
                    elem_senha.send_keys(self.credentials.senha + Keys.RETURN)
                else:
                    raise Exception("Campo senha nao encontrado")

                # Aguarda dashboard
                if self.selenium.aguardar_url_conter('dashboard'):
                    log.info("Login realizado com sucesso!")
                    return True

                raise Exception("Dashboard nao carregou apos login")

            except Exception as e:
                if attempt < max_retries:
                    wait_time = 2 ** attempt
                    log.warning(f"Login: tentativa {attempt + 1} falhou, retry em {wait_time}s... ({e})")
                    time.sleep(wait_time)
                else:
                    log.error(f"Login: falha apos {max_retries + 1} tentativas")
                    self.selenium.screenshot('erro_login')
                    return False

        return False

    def _processar_fundos_v14(self):
        """
        V14: Processa fundos com MULTIPLOS DRIVERS Chrome em paralelo.
        Usa ThreadPoolExecutor para criar N workers independentes.
        """
        fundos = list(self.fundo_manager.fundos.items())
        self.stats['total'] = len(fundos)

        # Determina tipo de relatorio
        tipo_download = None
        if self.flags.xml_enabled:
            tipo_download = 'XML'
        elif self.flags.excel_enabled:
            tipo_download = 'EXCEL'
        elif self.flags.pdf_enabled:
            tipo_download = 'PDF'

        if not tipo_download:
            log.warning("Nenhum tipo de download habilitado")
            return

        report_config = REPORT_CONFIGS.get(tipo_download)

        # =====================================================================
        # FASE 1: Coleta URLs e cookies no driver principal
        # =====================================================================
        print()
        log.info("=" * 50)
        log.info("FASE 1: Coletando URLs dos fundos...")
        log.info("=" * 50)

        self.selenium.navegar(self.credentials.url)
        self.selenium.aguardar_carregamento()

        # Coleta URLs
        fundos_urls = []
        for nome_fundo, pasta_fundo in fundos:
            sigla = self.fundo_manager.get_sigla(nome_fundo)
            try:
                elemento = self.selenium.encontrar_elemento(By.PARTIAL_LINK_TEXT, sigla)
                if elemento:
                    url = elemento.get_attribute('href')
                    fundos_urls.append({
                        'nome': nome_fundo,
                        'sigla': sigla,
                        'url': url,
                        'pasta': pasta_fundo
                    })
                    log.info(f"  URL: {sigla}")
                else:
                    log.warning(f"  {sigla} nao encontrado")
                    self.stats['erro'] += 1
            except Exception as e:
                log.warning(f"  Erro {sigla}: {e}")
                self.stats['erro'] += 1

        log.info(f"Total URLs coletadas: {len(fundos_urls)}")

        # Fecha driver principal (nao precisa mais)
        self.selenium.fechar()
        log.info("Driver principal fechado - workers farao login proprio")

        # =====================================================================
        # FASE 2: Processa fundos com multiplos workers
        # =====================================================================
        print()
        log.info("=" * 50)
        log.info(f"FASE 2: Iniciando {self.timeouts.NUM_WORKERS} workers (cada um faz login)...")
        log.info("=" * 50)

        resultados = []

        def worker_task(fundo_info: dict) -> dict:
            """Tarefa executada por cada worker."""
            worker_id = threading.current_thread().name[-1]

            # Cria worker Chrome com credenciais (fara login proprio)
            worker = WorkerChrome(
                worker_id=int(worker_id) if worker_id.isdigit() else 0,
                base_temp_path=self.paths.temp_download,
                credentials=self.credentials,
                datas=self.datas,
                report_config=report_config
            )

            try:
                if worker.iniciar():
                    result = worker.processar_fundo(
                        fundo_info['url'],
                        fundo_info['sigla'],
                        fundo_info['nome']
                    )
                else:
                    result = {
                        'nome': fundo_info['nome'],
                        'sigla': fundo_info['sigla'],
                        'status': 'erro'
                    }
            finally:
                worker.fechar()

            return result

        # Executa com ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=self.timeouts.NUM_WORKERS) as executor:
            futures = {
                executor.submit(worker_task, fundo): fundo
                for fundo in fundos_urls
            }

            for future in as_completed(futures):
                try:
                    result = future.result()
                    resultados.append(result)

                    if result['status'] == 'sucesso':
                        self.stats['sucesso'] += 1
                    else:
                        self.stats['erro'] += 1

                except Exception as e:
                    log.error(f"Worker falhou: {e}")
                    self.stats['erro'] += 1

        # =====================================================================
        # FASE 3: Processa arquivos baixados (de todas as pastas dos workers)
        # =====================================================================
        print()
        log.info("=" * 50)
        log.info("FASE 3: Processando arquivos...")
        log.info("=" * 50)

        # Busca ZIPs em todas as pastas dos workers (worker_0, worker_1, ...)
        zips = list(Path(self.paths.temp_download).glob('**/*.zip'))
        log.info(f"ZIPs encontrados: {len(zips)}")

        for zip_file in zips:
            self._processar_zip_v14(zip_file, tipo_download, fundos_urls)

        # Limpa pastas dos workers
        for worker_dir in Path(self.paths.temp_download).glob('worker_*'):
            try:
                shutil.rmtree(worker_dir, ignore_errors=True)
            except Exception:
                pass

    def _processar_zip_v14(self, zip_file: Path, report_type: str, fundos_urls: list):
        """Processa um ZIP baixado."""
        config = REPORT_CONFIGS.get(report_type.upper())
        if not config:
            return

        temp_extract = Path(self.file_handler.temp_path) / f'extract_{zip_file.stem}'
        temp_extract.mkdir(exist_ok=True)

        try:
            with zipfile.ZipFile(zip_file, 'r') as zf:
                zf.extractall(temp_extract)

            zip_file.unlink()

            for arquivo in temp_extract.glob(f'*{config.extension}'):
                nome_arquivo = arquivo.name.lower()

                for fundo_info in fundos_urls:
                    sigla = fundo_info['sigla'].lower()
                    nome_fundo = fundo_info['nome']
                    pasta = fundo_info.get('pasta', '')

                    if self.fundo_manager.is_bloko(nome_fundo):
                        padrao = self.fundo_manager.get_bloko_pattern(nome_fundo).lower()
                    else:
                        padrao = sigla

                    if padrao and padrao in nome_arquivo:
                        data_arquivo = extrair_data_de_nome_arquivo(arquivo.name)
                        if not data_arquivo:
                            data_arquivo = self.datas.data_final

                        self.file_handler._mover_arquivo(
                            arquivo, nome_fundo, data_arquivo, report_type, pasta
                        )
                        log.info(f"  Arquivo movido: {arquivo.name}")
                        break

        finally:
            if temp_extract.exists():
                shutil.rmtree(temp_extract, ignore_errors=True)

    def _processar_fundos(self):
        """
        DEPRECADO - Mantido para compatibilidade.
        Processa todos os fundos em PARALELO usando multiplas abas.
        Estrategia VERDADEIRAMENTE PARALELA:
        1. Coleta links de todos os fundos no dashboard
        2. Abre TODAS as guias de uma vez (29 guias)
        3. Em CADA guia, inicia o download (sem fechar)
        4. Aguarda TODOS os downloads completarem
        5. Fecha todas as guias extras
        6. Processa arquivos baixados
        """
        fundos = list(self.fundo_manager.fundos.items())

        # Determina qual tipo de relatorio baixar (prioridade: XML > Excel > PDF)
        tipo_download = None
        em_lote = False
        if self.flags.xml_enabled:
            tipo_download = 'XML'
            em_lote = self.flags.xml_lote
        elif self.flags.excel_enabled:
            tipo_download = 'EXCEL'
            em_lote = self.flags.excel_lote
        elif self.flags.pdf_enabled:
            tipo_download = 'PDF'
            em_lote = self.flags.pdf_lote

        if not tipo_download:
            log.warning("Nenhum tipo de download habilitado")
            return

        config = REPORT_CONFIGS.get(tipo_download)

        # Rastreia resultados e handles das abas
        resultados: Dict[str, str] = {}
        abas_fundos: Dict[str, dict] = {}  # {handle: {nome, sigla, pasta, url}}

        # =====================================================================
        # FASE 1: Coletar links dos fundos no dashboard
        # =====================================================================
        print()
        log.info("=" * 50)
        log.info("FASE 1: Coletando links dos fundos...")
        log.info("=" * 50)

        self.selenium.navegar(self.credentials.url)
        self.selenium.aguardar_carregamento()

        links_fundos: Dict[str, dict] = {}

        for nome_fundo, pasta_fundo in fundos:
            sigla = self.fundo_manager.get_sigla(nome_fundo)
            self.stats['total'] += 1

            try:
                elemento = self.selenium.encontrar_elemento(By.PARTIAL_LINK_TEXT, sigla)
                if elemento:
                    url = elemento.get_attribute('href')
                    links_fundos[nome_fundo] = {
                        'url': url,
                        'sigla': sigla,
                        'pasta': pasta_fundo
                    }
                    log.info(f"  Link encontrado: {sigla}")
                else:
                    log.warning(f"  {sigla} nao encontrado no dashboard")
                    resultados[nome_fundo] = 'erro'
            except Exception as e:
                log.warning(f"  Erro ao buscar {sigla}: {e}")
                resultados[nome_fundo] = 'erro'

        log.info(f"Total de links coletados: {len(links_fundos)}")

        # =====================================================================
        # FASE 2: Abrir TODAS as guias de uma vez
        # =====================================================================
        print()
        log.info("=" * 50)
        log.info("FASE 2: Abrindo TODAS as guias...")
        log.info("=" * 50)

        aba_principal = self.selenium.get_aba_principal()

        for idx, (nome_fundo, info) in enumerate(links_fundos.items(), 1):
            try:
                # Abre nova guia
                nova_aba = self.selenium.abrir_nova_aba()
                self.selenium.trocar_para_aba(nova_aba)

                # Navega para URL do fundo
                self.selenium.driver.get(info['url'])

                # Registra a aba
                abas_fundos[nova_aba] = {
                    'nome': nome_fundo,
                    'sigla': info['sigla'],
                    'pasta': info['pasta'],
                    'status': 'aberta'
                }

                log.info(f"  [{idx}/{len(links_fundos)}] Aba aberta: {info['sigla']}")

            except Exception as e:
                log.error(f"  Erro ao abrir aba para {info['sigla']}: {e}")
                resultados[nome_fundo] = 'erro'

        log.info(f"Total de abas abertas: {len(abas_fundos)}")

        # Volta para aba principal brevemente
        self.selenium.trocar_para_aba(aba_principal)

        # =====================================================================
        # FASE 3: Iniciar downloads RAPIDO em TODAS as guias
        # =====================================================================
        print()
        log.info("=" * 50)
        log.info("FASE 3: Disparando downloads (modo rapido)...")
        log.info("=" * 50)

        downloads_iniciados = 0

        for idx, (handle, info) in enumerate(abas_fundos.items(), 1):
            try:
                self.selenium.trocar_para_aba(handle)

                # Espera minima - apenas elemento clicavel
                time.sleep(0.3)

                log.info(f"  [{idx}/{len(abas_fundos)}] {info['sigla']}...")

                # Clica no botao do tipo de relatorio (rapido)
                if not self.selenium.clicar_elemento(
                    By.XPATH, f"//button[contains(., '{config.button_text}')]"
                ):
                    log.warning(f"    Botao nao encontrado")
                    info['status'] = 'erro'
                    continue

                time.sleep(0.3)

                # Inicia download RAPIDO (sem esperas desnecessarias)
                if em_lote and self._iniciar_download_rapido():
                    log.info(f"    OK!")
                    info['status'] = 'downloading'
                    downloads_iniciados += 1
                else:
                    log.warning(f"    Falha")
                    info['status'] = 'erro'

            except Exception as e:
                log.error(f"    Erro: {e}")
                info['status'] = 'erro'

        log.info(f"Downloads disparados: {downloads_iniciados}/{len(abas_fundos)}")

        # =====================================================================
        # FASE 4: Aguardar TODOS os downloads completarem
        # =====================================================================
        print()
        log.info("=" * 50)
        log.info(f"FASE 4: Aguardando {downloads_iniciados} downloads...")
        log.info("=" * 50)

        # Aguarda downloads completarem (polling)
        tempo_espera = max(downloads_iniciados * 3, 60)  # 3s por fundo, min 60s
        log.info(f"Aguardando ate {tempo_espera}s para downloads completarem...")

        for i in range(tempo_espera):
            time.sleep(1)
            # Verifica quantos ZIPs ja baixaram
            zips_baixados = len(list(Path(self.file_handler.temp_path).glob('*.zip')))
            if zips_baixados >= downloads_iniciados:
                log.info(f"  Todos os {zips_baixados} ZIPs baixados!")
                break
            if i % 10 == 0 and i > 0:
                log.info(f"  {i}s... {zips_baixados}/{downloads_iniciados} ZIPs")

        # =====================================================================
        # FASE 5: Fechar todas as guias extras
        # =====================================================================
        print()
        log.info("=" * 50)
        log.info("FASE 5: Fechando guias...")
        log.info("=" * 50)

        self.selenium.fechar_todas_abas_extras()
        log.info("  Todas as guias extras fechadas")

        # =====================================================================
        # FASE 6: Processar arquivos baixados
        # =====================================================================
        print()
        log.info("=" * 50)
        log.info("FASE 6: Processando arquivos...")
        log.info("=" * 50)

        # Lista todos os ZIPs
        zips = list(Path(self.file_handler.temp_path).glob('*.zip'))
        log.info(f"Encontrados {len(zips)} arquivos ZIP")

        for zip_file in zips:
            log.info(f"Extraindo: {zip_file.name}")
            self._processar_zip_v13(zip_file, tipo_download, abas_fundos)

        # =====================================================================
        # Contabilizar resultados
        # =====================================================================
        for handle, info in abas_fundos.items():
            if info['status'] in ['sucesso', 'downloading']:
                self.stats['sucesso'] += 1
            else:
                self.stats['erro'] += 1

        # Adiciona erros da fase de coleta de links
        self.stats['erro'] += len(resultados)

    def _processar_zip_v13(self, zip_file: Path, report_type: str,
                          abas_fundos: dict):
        """Processa um ZIP baixado, distribuindo arquivos para os fundos."""
        config = REPORT_CONFIGS.get(report_type.upper())
        if not config:
            return

        temp_extract = Path(self.file_handler.temp_path) / f'extract_{zip_file.stem}'
        temp_extract.mkdir(exist_ok=True)

        try:
            with zipfile.ZipFile(zip_file, 'r') as zf:
                zf.extractall(temp_extract)

            # Remove ZIP apos extrair
            zip_file.unlink()

            # Processa cada arquivo extraido
            for arquivo in temp_extract.glob(f'*{config.extension}'):
                nome_arquivo = arquivo.name.lower()

                # Tenta identificar qual fundo este arquivo pertence
                # Usa abas_fundos que tem {handle: {nome, sigla, pasta, status}}
                for _, info in abas_fundos.items():
                    sigla = info['sigla'].lower()
                    nome_fundo = info['nome']
                    pasta = info.get('pasta', '')

                    # Verifica se arquivo pertence a este fundo
                    if self.fundo_manager.is_bloko(nome_fundo):
                        padrao = self.fundo_manager.get_bloko_pattern(nome_fundo).lower()
                    else:
                        padrao = sigla

                    if padrao and padrao in nome_arquivo:
                        # Extrai data do arquivo
                        data_arquivo = extrair_data_de_nome_arquivo(arquivo.name)
                        if not data_arquivo:
                            data_arquivo = self.datas.data_final

                        # Move para destino
                        if self.file_handler._mover_arquivo(
                            arquivo, nome_fundo, data_arquivo, report_type, pasta
                        ):
                            info['status'] = 'sucesso'
                        break

        finally:
            if temp_extract.exists():
                shutil.rmtree(temp_extract, ignore_errors=True)

    def _iniciar_download_lote(self) -> bool:
        """Inicia download em lote na aba atual."""
        try:
            # Clica no menu de opcoes (...)
            menus = self.selenium.encontrar_elementos(
                By.XPATH, '//div[@data-kt-menu-trigger="click"]'
            )

            for menu in menus:
                try:
                    icone = menu.find_element(By.TAG_NAME, 'i')
                    if 'ellipsis' in icone.get_attribute('class'):
                        menu.click()
                        break
                except Exception:
                    continue

            self.selenium.aguardar_elemento_visivel(
                By.XPATH, '//a[contains(., "Download em Lote")]'
            )

            # Clica em "Download em Lote"
            if not self.selenium.clicar_elemento(
                By.XPATH, '//a[contains(., "Download em Lote")]'
            ):
                return False

            self.selenium.aguardar_elemento_visivel(By.ID, 'dataInicial')

            # Preenche datas
            self.selenium.preencher_data_js('dataInicial', self.datas.data_inicial)
            self.selenium.preencher_data_js('dataFinal', self.datas.data_final)

            self.selenium.aguardar_breve()

            # Clica em Download
            if not self.selenium.clicar_elemento(
                By.XPATH, '//button[contains(., "Download")]'
            ):
                return False

            return True

        except Exception as e:
            log.error(f"Erro ao iniciar download em lote: {e}")
            return False

    def _iniciar_download_rapido(self) -> bool:
        """
        Versao RAPIDA do download em lote - sem esperas desnecessarias.
        Usado no modo paralelo para disparar downloads rapidamente.
        """
        try:
            # Clica no menu de opcoes (...) usando JavaScript (mais rapido)
            self.selenium.driver.execute_script("""
                var menus = document.querySelectorAll('[data-kt-menu-trigger="click"]');
                for (var i = 0; i < menus.length; i++) {
                    var icone = menus[i].querySelector('i');
                    if (icone && icone.className.includes('ellipsis')) {
                        menus[i].click();
                        break;
                    }
                }
            """)

            time.sleep(0.2)

            # Clica em "Download em Lote" usando JavaScript
            self.selenium.driver.execute_script("""
                var links = document.querySelectorAll('a');
                for (var i = 0; i < links.length; i++) {
                    if (links[i].textContent.includes('Download em Lote')) {
                        links[i].click();
                        break;
                    }
                }
            """)

            time.sleep(0.3)

            # Preenche datas usando JavaScript (mais rapido)
            self.selenium.driver.execute_script(f"""
                var dataIni = document.getElementById('dataInicial');
                var dataFim = document.getElementById('dataFinal');
                if (dataIni) {{ dataIni.value = '{self.datas.data_inicial}'; }}
                if (dataFim) {{ dataFim.value = '{self.datas.data_final}'; }}
            """)

            time.sleep(0.2)

            # Clica em Download usando JavaScript
            self.selenium.driver.execute_script("""
                var botoes = document.querySelectorAll('button');
                for (var i = 0; i < botoes.length; i++) {
                    if (botoes[i].textContent.includes('Download')) {
                        botoes[i].click();
                        break;
                    }
                }
            """)

            return True

        except Exception as e:
            log.error(f"Erro download rapido: {e}")
            return False

    def _processar_zip_paralelo(self, zip_file: Path, report_type: str,
                                abas_fundos: dict):
        """Processa um ZIP baixado, distribuindo arquivos para os fundos."""
        config = REPORT_CONFIGS.get(report_type.upper())
        if not config:
            return

        temp_extract = Path(self.file_handler.temp_path) / f'extract_{zip_file.stem}'
        temp_extract.mkdir(exist_ok=True)

        try:
            with zipfile.ZipFile(zip_file, 'r') as zf:
                zf.extractall(temp_extract)

            # Remove ZIP apos extrair
            zip_file.unlink()

            # Processa cada arquivo extraido
            for arquivo in temp_extract.glob(f'*{config.extension}'):
                # Tenta identificar qual fundo este arquivo pertence
                nome_arquivo = arquivo.name.lower()

                for _, info in abas_fundos.items():
                    if info['status'] == 'erro':
                        continue

                    sigla = info['sigla'].lower()
                    nome = info['nome']
                    pasta = info.get('pasta', '')

                    # Verifica se arquivo pertence a este fundo
                    if self.fundo_manager.is_bloko(nome):
                        padrao = self.fundo_manager.get_bloko_pattern(nome).lower()
                    else:
                        padrao = sigla

                    if padrao and padrao in nome_arquivo:
                        # Extrai data do arquivo
                        data_arquivo = extrair_data_de_nome_arquivo(arquivo.name)
                        if not data_arquivo:
                            data_arquivo = self.datas.data_final

                        # Move para destino
                        if self.file_handler._mover_arquivo(
                            arquivo, nome, data_arquivo, report_type, pasta
                        ):
                            info['status'] = 'sucesso'
                        break

        finally:
            if temp_extract.exists():
                shutil.rmtree(temp_extract, ignore_errors=True)

    def _processar_fundo_com_retry(self, nome_fundo: str, sigla: str) -> bool:
        """Processa um fundo com retry automatico."""
        max_retries = self.timeouts.MAX_RETRIES

        for attempt in range(max_retries + 1):
            try:
                return self._processar_fundo_interno(nome_fundo, sigla)

            except Exception as e:
                if attempt < max_retries:
                    wait_time = 2 ** attempt
                    log.warning(f"Fundo {nome_fundo}: tentativa {attempt + 1} falhou, "
                               f"retry em {wait_time}s... ({e})")
                    time.sleep(wait_time)
                else:
                    log.error(f"Fundo {nome_fundo}: falha apos {max_retries + 1} tentativas - {e}")
                    return False

        return False

    def _sessao_valida(self) -> bool:
        """Verifica se a sessao ainda esta ativa (nao expirou)."""
        try:
            url_atual = self.selenium.driver.current_url.lower()
            # Se esta na pagina de login, sessao expirou
            return 'login' not in url_atual and 'signin' not in url_atual
        except Exception:
            return False

    def _verificar_e_recuperar_sessao(self) -> bool:
        """Verifica sessao e faz re-login se necessario."""
        if self._sessao_valida():
            return True

        log.warning("Sessao expirada, refazendo login...")
        return self._fazer_login()

    def _processar_fundo_interno(self, nome_fundo: str, sigla: str) -> bool:
        """Logica interna de processamento de um fundo."""
        # Verifica/recupera sessao antes de processar
        if not self._verificar_e_recuperar_sessao():
            raise Exception("Falha ao recuperar sessao")

        # Navega para dashboard
        self.selenium.navegar(self.credentials.url)

        # Clica no fundo
        if not self.selenium.clicar_elemento(By.PARTIAL_LINK_TEXT, sigla):
            raise Exception(f"Fundo {nome_fundo} nao encontrado no dashboard")

        # Aguarda carregar pagina do fundo
        self.selenium.aguardar_carregamento()

        sucesso_fundo = False

        # Processa cada tipo de relatorio
        if self.flags.pdf_enabled:
            if self._processar_tipo(nome_fundo, sigla, 'PDF', self.flags.pdf_lote):
                sucesso_fundo = True

        if self.flags.excel_enabled:
            # Re-navega para estado limpo
            self._renavegar_fundo(sigla)
            if self._processar_tipo(nome_fundo, sigla, 'EXCEL', self.flags.excel_lote):
                sucesso_fundo = True

        if self.flags.xml_enabled:
            self._renavegar_fundo(sigla)
            if self._processar_tipo(nome_fundo, sigla, 'XML', self.flags.xml_lote):
                sucesso_fundo = True

        return sucesso_fundo

    def _renavegar_fundo(self, sigla: str):
        """
        Re-navega para pagina do fundo de forma otimizada.
        Usa driver.back() quando possivel, evitando navegacao completa.
        """
        try:
            # Tenta usar back() - muito mais rapido
            self.selenium.driver.back()
            self.selenium.aguardar_carregamento()

            # Verifica se voltou para pagina do fundo (tem os botoes de relatorio)
            botoes = self.selenium.encontrar_elementos(
                By.XPATH, "//button[contains(., 'Carteira')]"
            )
            if botoes:
                return  # Sucesso - ja esta na pagina do fundo

        except Exception:
            pass

        # Fallback: navegacao completa (mais lenta, mas garantida)
        self.selenium.navegar(self.credentials.url)
        self.selenium.clicar_elemento(By.PARTIAL_LINK_TEXT, sigla)
        self.selenium.aguardar_carregamento()

    def _processar_tipo(self, fundo_nome: str, sigla: str,
                       report_type: str, em_lote: bool) -> bool:
        """Processa download de um tipo de relatorio."""
        config = REPORT_CONFIGS.get(report_type)
        if not config:
            return False

        log.info(f"  Baixando {report_type}...")

        # Clica no botao do tipo de relatorio
        if not self.selenium.clicar_elemento(
            By.XPATH, f"//button[contains(., '{config.button_text}')]"
        ):
            log.warning(f"  Botao '{config.button_text}' nao encontrado")
            return False

        self.selenium.aguardar_carregamento()

        if em_lote:
            return self._download_lote(fundo_nome, report_type, config)
        else:
            return self._download_individual(fundo_nome, sigla, report_type, config)

    def _download_individual(self, fundo_nome: str, sigla: str,
                            report_type: str, config: ReportConfig) -> bool:
        """Download de arquivo individual (modo nao-lote)."""
        data_busca = self.datas.data_exibicao

        # Procura na tabela
        rows = self.selenium.encontrar_elementos(By.XPATH, '//tbody/tr')

        for row in rows:
            try:
                texto = row.text
                if data_busca in texto:
                    # Encontrou a linha, clica no botao de download
                    btn = row.find_element(By.TAG_NAME, 'button')
                    btn.click()

                    log.info(f"  Download iniciado para {data_busca}")

                    # Aguarda arquivo
                    arquivo = self.file_handler.aguardar_download(
                        config.extension, sigla
                    )

                    if arquivo:
                        pasta_fundo = self.fundo_manager.fundos.get(fundo_nome, '')
                        return self.file_handler.processar_arquivo_individual(
                            arquivo, fundo_nome, self.datas.data_inicial,
                            report_type, pasta_fundo
                        )
                    else:
                        log.warning("  Timeout aguardando download")
                        return False

            except Exception:
                continue

        log.warning(f"  Data {data_busca} nao encontrada na tabela")
        return False

    def _download_lote(self, fundo_nome: str, report_type: str,
                      config: ReportConfig) -> bool:
        """Download em lote (ZIP)."""
        try:
            # Clica no menu de opcoes (...)
            menus = self.selenium.encontrar_elementos(
                By.XPATH, '//div[@data-kt-menu-trigger="click"]'
            )

            for menu in menus:
                try:
                    icone = menu.find_element(By.TAG_NAME, 'i')
                    if 'ellipsis' in icone.get_attribute('class'):
                        menu.click()
                        break
                except Exception:
                    continue

            self.selenium.aguardar_elemento_visivel(By.XPATH, '//a[contains(., "Download em Lote")]')

            # Clica em "Download em Lote"
            if not self.selenium.clicar_elemento(
                By.XPATH, '//a[contains(., "Download em Lote")]'
            ):
                log.warning("  Opcao 'Download em Lote' nao encontrada")
                return False

            self.selenium.aguardar_elemento_visivel(By.ID, 'dataInicial')

            # Preenche datas
            self.selenium.preencher_data_js('dataInicial', self.datas.data_inicial)
            self.selenium.preencher_data_js('dataFinal', self.datas.data_final)

            self.selenium.aguardar_breve()

            # Clica em Download
            if not self.selenium.clicar_elemento(
                By.XPATH, '//button[contains(., "Download")]'
            ):
                return False

            log.info("  Download em lote iniciado...")

            # Limpa pasta temp antes de aguardar novo ZIP
            self.file_handler.limpar_temp()

            # Aguarda o ZIP terminar de baixar
            zip_baixado = self.file_handler.aguardar_download('.zip', timeout=60)

            if not zip_baixado:
                log.warning("  Timeout aguardando download do ZIP")
                return False

            log.info(f"  ZIP baixado: {zip_baixado.name}")

            # Processa ZIP
            qtd = self.file_handler.processar_zip_lote(
                fundo_nome, self.fundo_manager, report_type,
                self.datas.data_final
            )

            if qtd > 0:
                log.info(f"  {qtd} arquivo(s) processado(s)")
                return True
            else:
                log.warning("  Nenhum arquivo processado do ZIP")
                return False

        except Exception as e:
            log.error(f"  Erro no download em lote: {e}")
            return False


# =============================================================================
# FUNCAO DE CARREGAMENTO DE CONFIGURACAO
# =============================================================================

def carregar_config_planilha(caminho: str) -> Tuple[QorePaths, QoreCredentials, QoreFlags, QoreDatas]:
    """Carrega configuracoes da planilha DOWNLOADS_AUX.xlsx."""

    log.info(f"Lendo configuracoes de: {caminho}")

    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb['Downloads']

    # Datas
    data_ini = ws['C4'].value
    data_fim = ws['C5'].value

    if isinstance(data_ini, str):
        data_ini = datetime.strptime(data_ini, '%d/%m/%Y')
    if isinstance(data_fim, str):
        data_fim = datetime.strptime(data_fim, '%d/%m/%Y')
    if data_fim is None:
        data_fim = data_ini

    # Paths
    paths = QorePaths(
        pdf=str(ws['I9'].value or '').strip(),
        excel=str(ws['I13'].value or '').strip(),
        xml=str(ws['I21'].value or r'C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\03. Arquivos Rotina\XML_QORE').strip(),
        pdf_monitoramento=r'C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\03. Arquivos Rotina\05. PDF',
        temp_download=str(ws['I20'].value or '').strip(),
        bd_path=str(ws['I19'].value or '').strip()
    )

    # Credenciais
    credentials = QoreCredentials(
        url=str(ws['M10'].value or '').strip(),
        email=str(ws['N10'].value or '').strip(),
        senha=str(ws['O10'].value or '').strip()
    )

    # Flags
    flags = QoreFlags(
        qore_enabled=validar_boolean(ws['C24'].value),
        pdf_enabled=validar_boolean(ws['C25'].value),
        pdf_lote=validar_boolean(ws['C26'].value),
        excel_enabled=validar_boolean(ws['C27'].value),
        excel_lote=validar_boolean(ws['C28'].value),
        xml_enabled=validar_boolean(ws['C29'].value),
        xml_lote=validar_boolean(ws['C30'].value)
    )

    # Datas
    datas = QoreDatas(data_inicial=data_ini, data_final=data_fim)

    return paths, credentials, flags, datas


# =============================================================================
# MAIN
# =============================================================================

if __name__ == '__main__':
    DEFAULT_AUX_PATH = r'C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\03. Arquivos Rotina\07. DEPARA\DOWNLOADS_AUX.xlsx'

    # Caminho da planilha
    aux_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_AUX_PATH

    if not os.path.exists(aux_path):
        log.critical(f"Planilha nao encontrada: {aux_path}")
        sys.exit(1)

    try:
        # Carrega configuracoes
        paths, credentials, flags, datas = carregar_config_planilha(aux_path)

        # Verifica se esta habilitado
        if not flags.qore_enabled:
            log.info("Automacao QORE desabilitada na planilha (C24)")
            sys.exit(0)

        # Verifica se tem algo para baixar
        if not any([flags.pdf_enabled, flags.excel_enabled, flags.xml_enabled]):
            log.info("Nenhum tipo de download habilitado")
            sys.exit(0)

        # Cria pastas necessarias
        os.makedirs(paths.temp_download, exist_ok=True)
        os.makedirs(paths.pdf_monitoramento, exist_ok=True)
        if paths.xml:
            os.makedirs(paths.xml, exist_ok=True)

        # Executa
        bot = QoreAutomation(paths, credentials, flags, datas)
        bot.executar()

    except Exception as e:
        log.critical(f"Erro fatal: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
