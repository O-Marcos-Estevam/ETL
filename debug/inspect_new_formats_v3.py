import os
import openpyxl

base_path = r'C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\02. Dev\ETL\Relatórios\Novo'
excel_name = 'CARTEIRA_DIARIA_55523261_08-12-2025-5d6909a2-64d7-4af2-8ecd-0dba2748e744.xlsx'
xml_name = 'XML5.xml'

print(f"Base Path: {base_path}")

print("\n--- INSPECTING XML HEAD ---")
xml_path = os.path.join(base_path, xml_name)
if os.path.exists(xml_path):
    try:
        with open(xml_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            print(f"First 500 chars:\n{content[:500]}")
    except Exception as e:
        print(f"Error reading XML text: {e}")
else:
    print(f"XML not found at {xml_path}")

print("\n--- INSPECTING EXCEL STRUCTURE ---")
excel_path = os.path.join(base_path, excel_name)
if os.path.exists(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        print(f"Sheet names: {wb.sheetnames}")
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            print(f"-- Sheet: {sheet} --")
            for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
                print(f"Row {i}: {row}")
    except Exception as e:
        print(f"Error reading Excel: {e}")
else:
    print(f"Excel not found at {excel_path}")
