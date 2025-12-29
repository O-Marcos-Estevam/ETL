"""
================================================================================
CAPTURA QORE CDP - Usando Chrome DevTools Protocol
================================================================================

Versao que usa CDP nativo do Selenium para capturar chamadas de rede.
Nao requer selenium-wire.

Autor: ETL Team
Data: Dezembro 2025
================================================================================
"""

import os
import sys
import json
import time
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any
from urllib.parse import urlparse

sys.path.insert(0, str(Path(__file__).parent.parent))

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

import openpyxl


class CapturaQoreCDP:
    """Captura informacoes do QORE usando Chrome DevTools Protocol."""

    def __init__(self, output_dir: str = "dump_qore"):
        self.output_path = Path(output_dir)
        self.driver = None

        # Dados capturados
        self.network_requests: List[Dict] = []
        self.fundos: List[Dict] = []
        self.cookies: List[Dict] = []
        self.endpoints_unicos: Dict[str, Dict] = {}

    def _log(self, msg: str, level: str = "INFO"):
        symbols = {"INFO": "[+]", "WARNING": "[!]", "ERROR": "[X]", "DEBUG": "[.]"}
        print(f"{symbols.get(level, '[?]')} {msg}")

    def _criar_estrutura(self):
        """Cria estrutura de pastas."""
        for pasta in ["html", "api", "screenshots", "fundos"]:
            (self.output_path / pasta).mkdir(parents=True, exist_ok=True)
        self._log(f"Output em: {self.output_path.absolute()}")

    def _iniciar_driver(self):
        """Inicia Chrome com logging de rede habilitado."""
        self._log("Iniciando Chrome com CDP...")

        chrome_options = Options()
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--start-maximized')
        chrome_options.add_argument('--disable-extensions')

        # Habilita logging de performance (inclui Network)
        chrome_options.set_capability('goog:loggingPrefs', {'performance': 'ALL'})

        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.implicitly_wait(5)

        # Habilita Network domain via CDP
        self.driver.execute_cdp_cmd('Network.enable', {})

        self._log("Chrome iniciado com captura de rede")

    def _extrair_logs_rede(self) -> List[Dict]:
        """Extrai logs de rede do Chrome."""
        logs = []

        try:
            perf_logs = self.driver.get_log('performance')

            for entry in perf_logs:
                try:
                    log = json.loads(entry['message'])
                    message = log.get('message', {})
                    method = message.get('method', '')
                    params = message.get('params', {})

                    # Captura requests
                    if method == 'Network.requestWillBeSent':
                        request = params.get('request', {})
                        url = request.get('url', '')

                        logs.append({
                            'type': 'request',
                            'requestId': params.get('requestId'),
                            'method': request.get('method'),
                            'url': url,
                            'headers': request.get('headers', {}),
                            'postData': request.get('postData'),
                            'timestamp': params.get('timestamp')
                        })

                    # Captura responses
                    elif method == 'Network.responseReceived':
                        response = params.get('response', {})
                        url = response.get('url', '')

                        logs.append({
                            'type': 'response',
                            'requestId': params.get('requestId'),
                            'url': url,
                            'status': response.get('status'),
                            'mimeType': response.get('mimeType'),
                            'headers': response.get('headers', {}),
                            'timestamp': params.get('timestamp')
                        })

                except Exception:
                    continue

        except Exception as e:
            self._log(f"Erro ao extrair logs: {e}", "DEBUG")

        return logs

    def _processar_logs_rede(self):
        """Processa logs de rede e identifica endpoints."""
        logs = self._extrair_logs_rede()

        for log in logs:
            url = log.get('url', '')

            # Filtra assets estaticos
            is_asset = any(ext in url.lower() for ext in [
                '.js', '.css', '.png', '.jpg', '.jpeg', '.gif',
                '.svg', '.woff', '.woff2', '.ttf', '.ico', '.map'
            ])

            # Identifica APIs
            is_api = any(x in url.lower() for x in [
                '/api/', 'graphql', '/v1/', '/v2/',
                'fundos', 'boletas', 'carteira', 'download',
                'relatorio', 'posicao', 'cota', 'patrimonio',
                'portfolios'
            ])

            if is_api or (not is_asset and log.get('type') == 'request'):
                self.network_requests.append(log)

                # Endpoint unico
                parsed = urlparse(url)
                method = log.get('method', 'GET')
                endpoint_key = f"{method} {parsed.path}"

                if endpoint_key not in self.endpoints_unicos and parsed.path:
                    self.endpoints_unicos[endpoint_key] = {
                        'method': method,
                        'path': parsed.path,
                        'url': url,
                        'status': log.get('status'),
                        'mimeType': log.get('mimeType'),
                        'headers': log.get('headers', {})
                    }
                    self._log(f"  Endpoint: {endpoint_key}", "DEBUG")

    def _fazer_login(self, url: str, email: str, senha: str) -> bool:
        """Realiza login no QORE."""
        try:
            self._log(f"Acessando: {url}")
            self.driver.get(url)
            time.sleep(2)

            # Captura pagina de login
            self._salvar_html("01_login")
            self._screenshot("01_login")

            # Login
            self._log("Inserindo credenciais...")

            email_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.NAME, 'email'))
            )
            email_field.clear()
            email_field.send_keys(email)

            senha_field = self.driver.find_element(By.NAME, 'password')
            senha_field.clear()
            senha_field.send_keys(senha + Keys.RETURN)

            # Aguarda dashboard
            WebDriverWait(self.driver, 15).until(
                lambda d: 'dashboard' in d.current_url.lower()
            )

            self._log("Login OK!")
            time.sleep(2)

            # Processa logs de rede do login
            self._processar_logs_rede()

            # Captura cookies
            self.cookies = self.driver.get_cookies()
            self._salvar_json("cookies.json", self.cookies)
            self._log(f"Capturados {len(self.cookies)} cookies")

            return True

        except Exception as e:
            self._log(f"Erro no login: {e}", "ERROR")
            self._screenshot("erro_login")
            return False

    def _capturar_dashboard(self):
        """Captura dashboard e lista de fundos."""
        self._log("Capturando dashboard...")

        self._salvar_html("02_dashboard")
        self._screenshot("02_dashboard")

        # Processa logs de rede
        self._processar_logs_rede()

        # Extrai fundos
        self._extrair_fundos()

    def _extrair_fundos(self):
        """Extrai lista de fundos do dashboard."""
        self._log("Extraindo lista de fundos...")

        try:
            links = self.driver.find_elements(
                By.XPATH,
                "//a[contains(@href, 'portfolios') or contains(@href, 'fundo')]"
            )

            urls_vistas = set()

            for link in links:
                try:
                    href = link.get_attribute('href')
                    texto = link.text.strip()

                    if href and texto and href not in urls_vistas:
                        urls_vistas.add(href)
                        self.fundos.append({
                            'nome': texto,
                            'url': href
                        })
                except:
                    continue

            self._log(f"  {len(self.fundos)} fundos encontrados")
            self._salvar_json("fundos/lista_fundos.json", self.fundos)

        except Exception as e:
            self._log(f"Erro ao extrair fundos: {e}", "ERROR")

    def _capturar_fundo(self, fundo: Dict, indice: int):
        """Captura pagina de um fundo."""
        nome = fundo['nome']
        url = fundo['url']

        self._log(f"Capturando fundo {indice + 1}: {nome}")

        try:
            self.driver.get(url)
            time.sleep(1.5)

            nome_arquivo = f"fundo_{indice + 1:02d}_{self._sanitizar(nome)}"
            self._salvar_html(nome_arquivo)
            self._screenshot(nome_arquivo)

            # Processa logs de rede
            self._processar_logs_rede()

            # Tenta interagir com botoes de relatorio
            self._explorar_botoes_relatorio()

        except Exception as e:
            self._log(f"Erro ao capturar {nome}: {e}", "ERROR")

    def _explorar_botoes_relatorio(self):
        """Explora botoes de relatorio para capturar endpoints."""
        try:
            # Clica em XML Anbima 5.0 (mais interessante)
            botoes = self.driver.find_elements(
                By.XPATH,
                "//button[contains(., 'XML') or contains(., 'Excel') or contains(., 'PDF')]"
            )

            if botoes:
                self._log(f"  Botoes: {[b.text for b in botoes[:4]]}", "DEBUG")

                # Clica no primeiro botao (geralmente XML)
                botoes[0].click()
                time.sleep(1)

                # Processa logs
                self._processar_logs_rede()

                # Tenta abrir menu de download em lote
                self._explorar_download_lote()

        except Exception as e:
            self._log(f"  Erro ao explorar botoes: {e}", "DEBUG")

    def _explorar_download_lote(self):
        """Explora modal de download em lote."""
        try:
            # Procura menu (...)
            menus = self.driver.find_elements(
                By.XPATH,
                '//div[@data-kt-menu-trigger="click"]'
            )

            for menu in menus:
                try:
                    icone = menu.find_element(By.TAG_NAME, 'i')
                    if 'ellipsis' in icone.get_attribute('class'):
                        menu.click()
                        time.sleep(0.5)

                        # Processa logs
                        self._processar_logs_rede()

                        # Clica em Download em Lote
                        link_lote = self.driver.find_element(
                            By.XPATH,
                            '//a[contains(., "Download em Lote")]'
                        )
                        link_lote.click()
                        time.sleep(0.5)

                        # Captura modal
                        self._salvar_html("modal_download_lote")
                        self._screenshot("modal_download_lote")

                        # Processa logs
                        self._processar_logs_rede()

                        # Fecha modal
                        self.driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
                        time.sleep(0.3)

                        break
                except:
                    continue

        except Exception as e:
            self._log(f"  Erro ao explorar download lote: {e}", "DEBUG")

    def _salvar_html(self, nome: str):
        """Salva HTML da pagina atual."""
        arquivo = self.output_path / "html" / f"{nome}.html"
        arquivo.write_text(self.driver.page_source, encoding='utf-8')

    def _screenshot(self, nome: str):
        """Salva screenshot."""
        arquivo = self.output_path / "screenshots" / f"{nome}.png"
        try:
            self.driver.save_screenshot(str(arquivo))
        except:
            pass

    def _salvar_json(self, nome: str, dados: Any):
        """Salva dados em JSON."""
        arquivo = self.output_path / nome
        arquivo.parent.mkdir(parents=True, exist_ok=True)
        with open(arquivo, 'w', encoding='utf-8') as f:
            json.dump(dados, f, indent=2, ensure_ascii=False, default=str)

    def _sanitizar(self, nome: str) -> str:
        """Sanitiza nome para arquivo."""
        return re.sub(r'[^\w\-]', '_', nome)[:50]

    def _gerar_relatorio(self):
        """Gera relatorio final."""
        self._log("Gerando relatorio...")

        # Salva endpoints
        self._salvar_json("api/endpoints_unicos.json", self.endpoints_unicos)
        self._salvar_json("api/todas_requests.json", self.network_requests)

        # Relatorio
        relatorio = {
            'data_captura': datetime.now().isoformat(),
            'total_fundos': len(self.fundos),
            'total_requests': len(self.network_requests),
            'total_endpoints': len(self.endpoints_unicos),
            'endpoints': list(self.endpoints_unicos.keys())
        }
        self._salvar_json("RELATORIO.json", relatorio)

        # Imprime resumo
        print("\n" + "=" * 70)
        print("  RELATORIO DE CAPTURA CDP")
        print("=" * 70)
        print(f"  Fundos encontrados:    {len(self.fundos)}")
        print(f"  Requests capturadas:   {len(self.network_requests)}")
        print(f"  Endpoints unicos:      {len(self.endpoints_unicos)}")
        print("=" * 70)

        if self.endpoints_unicos:
            print("\n  ENDPOINTS ENCONTRADOS:")
            for ep in list(self.endpoints_unicos.keys())[:20]:
                print(f"    - {ep}")
            if len(self.endpoints_unicos) > 20:
                print(f"    ... e mais {len(self.endpoints_unicos) - 20}")

        print(f"\n  Output salvo em: {self.output_path.absolute()}")
        print("=" * 70)

    def executar(self, url: str, email: str, senha: str, max_fundos: int = 3):
        """Executa captura completa."""
        print("\n" + "=" * 70)
        print("  CAPTURA QORE - Chrome DevTools Protocol")
        print("=" * 70 + "\n")

        self._criar_estrutura()

        try:
            self._iniciar_driver()

            if not self._fazer_login(url, email, senha):
                return

            self._capturar_dashboard()

            # Captura alguns fundos
            for i, fundo in enumerate(self.fundos[:max_fundos]):
                self._capturar_fundo(fundo, i)

            self._gerar_relatorio()

        except Exception as e:
            self._log(f"Erro fatal: {e}", "ERROR")
            import traceback
            traceback.print_exc()

        finally:
            if self.driver:
                self.driver.quit()
                self._log("Chrome fechado")


def carregar_credenciais(caminho: str):
    """Carrega credenciais da planilha."""
    try:
        wb = openpyxl.load_workbook(caminho, data_only=True)
        ws = wb['Downloads']
        return (
            str(ws['M10'].value or '').strip(),
            str(ws['N10'].value or '').strip(),
            str(ws['O10'].value or '').strip()
        )
    except Exception as e:
        print(f"[X] Erro ao ler planilha: {e}")
        return None, None, None


if __name__ == '__main__':
    PLANILHA = r'C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\03. Arquivos Rotina\07. DEPARA\DOWNLOADS_AUX.xlsx'

    url, email, senha = carregar_credenciais(PLANILHA)

    if not all([url, email, senha]):
        print("\n[!] Configure credenciais manualmente:")
        url = input("URL: ").strip() or "https://hub.qoredtvm.com.br"
        email = input("Email: ").strip()
        senha = input("Senha: ").strip()

    if not all([email, senha]):
        print("[X] Credenciais invalidas")
        sys.exit(1)

    captura = CapturaQoreCDP(output_dir="dump_qore")
    captura.executar(url, email, senha, max_fundos=3)
