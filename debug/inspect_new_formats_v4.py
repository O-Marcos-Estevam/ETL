import os
import openpyxl

base_path = r'C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\02. Dev\ETL\Relatórios\Novo'
excel_name = 'CARTEIRA_DIARIA_55523261_08-12-2025-5d6909a2-64d7-4af2-8ecd-0dba2748e744.xlsx'
xml_name = 'XML5.xml'
output_file = r'C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\02. Dev\ETL\inspection_result.txt'

with open(output_file, 'w', encoding='utf-8') as out:
    out.write(f"Base Path: {base_path}\n")

    out.write("\n--- INSPECTING XML HEAD ---\n")
    xml_path = os.path.join(base_path, xml_name)
    if os.path.exists(xml_path):
        try:
            with open(xml_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                out.write(f"First 1000 chars:\n{content[:1000]}\n")
        except Exception as e:
            out.write(f"Error reading XML text: {e}\n")
    else:
        out.write(f"XML not found at {xml_path}\n")

    out.write("\n--- INSPECTING EXCEL STRUCTURE ---\n")
    excel_path = os.path.join(base_path, excel_name)
    if os.path.exists(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            out.write(f"Sheet names: {wb.sheetnames}\n")
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                out.write(f"-- Sheet: {sheet} --\n")
                for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
                    out.write(f"Row {i}: {row}\n")
        except Exception as e:
            out.write(f"Error reading Excel: {e}\n")
    else:
        out.write(f"Excel not found at {excel_path}\n")

print("Done writing to " + output_file)
