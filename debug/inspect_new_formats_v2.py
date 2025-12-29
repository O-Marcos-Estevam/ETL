import os
import zipfile
from xml.etree import ElementTree as ET

base_path = r'C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\02. Dev\ETL'
excel_name = 'CARTEIRA_DIARIA_55523261_08-12-2025-5d6909a2-64d7-4af2-8ecd-0dba2748e744.xlsx'
xml_name = 'XML5.xml'

print("--- FILE CHECK ---")
excel_path = os.path.join(base_path, excel_name)
xml_path = os.path.join(base_path, xml_name)
print(f"Excel exists? {os.path.exists(excel_path)}")
print(f"XML exists? {os.path.exists(xml_path)}")

print("\n--- INSPECTING XML HEAD ---")
if os.path.exists(xml_path):
    try:
        with open(xml_path, 'r', encoding='utf-8', errors='ignore') as f:
            print(f.read()[:500])
    except Exception as e:
        print(f"Error reading XML text: {e}")

print("\n--- INSPECTING EXCEL STRUCTURE (OPENPYXL) ---")
if os.path.exists(excel_path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        print(f"Sheet names: {wb.sheetnames}")
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            print(f"-- Sheet: {sheet} --")
            # Print first 5 rows
            for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
                print(f"Row {i}: {row}")
    except Exception as e:
        print(f"Error reading Excel with openpyxl: {e}")
