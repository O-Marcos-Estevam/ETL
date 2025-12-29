"""
================================================================================
TESTE DE API QORE - Descoberta de Endpoints
================================================================================

Script para testar autenticacao e descobrir os parametros corretos da API.
Ajuda a identificar os valores de 'tipo' para Excel e XML.

Uso: python debug/test_api_qore.py
================================================================================
"""

import os
import sys
import json
from pathlib import Path
from datetime import datetime

# Adiciona path do projeto
sys.path.insert(0, str(Path(__file__).parent.parent))

import requests
import openpyxl


def carregar_credenciais(planilha_path: str):
    """Carrega credenciais da planilha."""
    try:
        wb = openpyxl.load_workbook(planilha_path, data_only=True)
        ws = wb['Downloads']
        return {
            'url': str(ws['M10'].value or '').strip(),
            'email': str(ws['N10'].value or '').strip(),
            'senha': str(ws['O10'].value or '').strip()
        }
    except Exception as e:
        print(f"[X] Erro ao carregar credenciais: {e}")
        return None


def autenticar(credentials: dict) -> str:
    """Autentica e retorna o token JWT."""
    print("\n[1] AUTENTICACAO")
    print("=" * 50)

    url = 'https://hub.qoredtvm.com.br/api/v1/authorize'

    payload = {
        'username': credentials['email'],
        'password': credentials['senha']
    }

    print(f"[+] URL: {url}")
    print(f"[+] Email: {credentials['email']}")

    try:
        response = requests.post(url, json=payload, timeout=10)

        print(f"[+] Status: {response.status_code}")

        if response.status_code == 200:
            data = response.json()
            print(f"[+] Resposta: {json.dumps(data, indent=2)[:500]}...")

            token = data.get('access_token') or data.get('token') or data.get('accessToken')
            if token:
                print(f"[+] Token obtido: {token[:50]}...")
                return token
            else:
                print("[X] Token nao encontrado na resposta")
                print(f"    Chaves disponiveis: {list(data.keys())}")
        else:
            print(f"[X] Falha: {response.text}")

    except Exception as e:
        print(f"[X] Erro: {e}")

    return None


def listar_arquivos_fundo(token: str, uuid: str, tipo: str):
    """Lista arquivos de um fundo."""
    url = f'https://hub.qoredtvm.com.br/api/v1/fundos-posicao/{uuid}/arquivos'

    headers = {
        'Authorization': f'Bearer {token}',
        'Accept': 'application/json'
    }

    params = {
        'tipo': tipo,
        'p': 0
    }

    try:
        response = requests.get(url, headers=headers, params=params, timeout=10)
        return response.status_code, response.json() if response.status_code == 200 else response.text

    except Exception as e:
        return None, str(e)


def testar_tipos_arquivo(token: str, uuid: str):
    """Testa diferentes valores de 'tipo' para descobrir os corretos."""
    print("\n[2] DESCOBERTA DE TIPOS DE ARQUIVO")
    print("=" * 50)

    # Tipos conhecidos e confirmados via captura de endpoint
    # Descoberto via captura_xml_endpoint.py usando Chrome DevTools Protocol
    tipos_testar = [
        # PDF (confirmado)
        'CARTEIRA_PDF',
        # Excel (confirmado)
        'CARTEIRA_EXCEL',
        # XML Anbima 5.0 (CONFIRMADO via captura)
        'XML_5_0',
        # XML Anbima 4.01 (provavel - mesmo padrao)
        'XML_4_01',
        'XML_4_0',
    ]

    print(f"[+] UUID de teste: {uuid}")
    print(f"[+] Testando {len(tipos_testar)} valores de 'tipo'...")
    print()

    tipos_validos = []

    for tipo in tipos_testar:
        status, data = listar_arquivos_fundo(token, uuid, tipo)

        if status == 200:
            qtd = len(data) if isinstance(data, list) else len(data.get('content', []))
            print(f"  [OK] {tipo:25s} -> {qtd} arquivos")
            tipos_validos.append({'tipo': tipo, 'qtd': qtd})
        else:
            print(f"  [--] {tipo:25s} -> {status}")

    print()
    print("[+] TIPOS VALIDOS ENCONTRADOS:")
    for t in tipos_validos:
        print(f"    - {t['tipo']}: {t['qtd']} arquivos")

    return tipos_validos


def analisar_estrutura_arquivos(token: str, uuid: str, tipo: str):
    """Analisa estrutura da resposta de listagem de arquivos."""
    print(f"\n[3] ESTRUTURA DA LISTAGEM ({tipo})")
    print("=" * 50)

    status, data = listar_arquivos_fundo(token, uuid, tipo)

    if status == 200:
        # Verifica se e lista ou dict
        if isinstance(data, dict):
            print(f"[+] Resposta e dict com chaves: {list(data.keys())}")
            arquivos = data.get('content', data.get('items', data.get('data', [])))
        else:
            arquivos = data
            print(f"[+] Resposta e lista com {len(arquivos)} itens")

        if arquivos and len(arquivos) > 0:
            primeiro = arquivos[0]
            print(f"\n[+] Estrutura do primeiro arquivo:")
            print(f"    Chaves: {list(primeiro.keys())}")
            print(f"\n[+] Valores:")
            for k, v in primeiro.items():
                valor_str = str(v)[:100] if v else 'None'
                print(f"    {k}: {valor_str}")

            return primeiro
    else:
        print(f"[X] Erro ao listar: {status}")

    return None


def testar_download_individual(token: str, uuid: str, arquivo_info: dict):
    """Testa download de arquivo individual."""
    print(f"\n[4] TESTE DE DOWNLOAD INDIVIDUAL")
    print("=" * 50)

    headers = {
        'Authorization': f'Bearer {token}',
        'Accept': 'application/json, application/octet-stream, application/pdf, */*'
    }

    # Tenta encontrar URL de download no arquivo_info
    possiveis_campos = ['url', 'downloadUrl', 'link', 'arquivo', 'path', 'uri', 'href']

    download_url = None
    for campo in possiveis_campos:
        if campo in arquivo_info and arquivo_info[campo]:
            download_url = arquivo_info[campo]
            print(f"[+] Campo de download encontrado: {campo}")
            break

    if not download_url:
        print("[!] URL de download nao encontrada nos campos conhecidos")

        # Tenta construir URL baseada no guid/id
        arquivo_guid = arquivo_info.get('guid') or arquivo_info.get('id') or arquivo_info.get('arquivoId')
        if arquivo_guid:
            print(f"[+] Tentando com GUID do arquivo: {arquivo_guid}")

            # Endpoints possiveis para download individual
            endpoints = [
                f'/api/v1/arquivos/{arquivo_guid}/download',
                f'/api/v1/fundos-posicao/{uuid}/arquivos/{arquivo_guid}/download',
                f'/api/v1/fundos-posicao/{uuid}/arquivos/{arquivo_guid}',
                f'/api/v1/download/{arquivo_guid}',
            ]

            for endpoint in endpoints:
                url = f'https://hub.qoredtvm.com.br{endpoint}'

                try:
                    response = requests.get(url, headers=headers, timeout=30, stream=True)

                    content_type = response.headers.get('content-type', '')
                    print(f"\n  Tentando: {endpoint}")
                    print(f"  Status: {response.status_code}, Content-Type: {content_type}")

                    if response.status_code == 200 and ('pdf' in content_type or 'octet' in content_type or 'excel' in content_type):
                        print(f"  [OK] Download disponivel!")

                        # Salva arquivo de teste
                        test_file = Path(__file__).parent / 'test_download.pdf'
                        with open(test_file, 'wb') as f:
                            for chunk in response.iter_content(8192):
                                f.write(chunk)
                        print(f"  [OK] Arquivo salvo: {test_file}")
                        print(f"  [OK] Tamanho: {test_file.stat().st_size} bytes")
                        return endpoint

                except Exception as e:
                    print(f"  [X] Erro: {e}")

    else:
        # Tenta download direto
        if not download_url.startswith('http'):
            download_url = f'https://hub.qoredtvm.com.br{download_url}'

        print(f"[+] Tentando URL: {download_url}")

        try:
            response = requests.get(download_url, headers=headers, timeout=30, stream=True)

            print(f"  Status: {response.status_code}")
            print(f"  Content-Type: {response.headers.get('content-type', '')}")

            if response.status_code == 200:
                print(f"  [OK] Download funcionou!")
                return download_url

        except Exception as e:
            print(f"  [X] Erro: {e}")

    return None


def testar_download(token: str, uuid: str, tipo: str):
    """Testa endpoints de download em lote."""
    print(f"\n[5] TESTE DE DOWNLOAD EM LOTE ({tipo})")
    print("=" * 50)

    headers = {
        'Authorization': f'Bearer {token}',
        'Accept': 'application/json, application/octet-stream, */*'
    }

    # Endpoints de download possiveis
    endpoints = [
        f'/api/v1/fundos-posicao/{uuid}/arquivos/download?tipo={tipo}',
        f'/api/v1/fundos-posicao/{uuid}/arquivos/download?tipo={tipo}&dataInicial=2025-12-01&dataFinal=2025-12-28',
        f'/api/v1/fundos-posicao/{uuid}/download?tipo={tipo}',
    ]

    for endpoint in endpoints:
        url = f'https://hub.qoredtvm.com.br{endpoint}'

        try:
            response = requests.get(url, headers=headers, timeout=30, stream=True)

            content_type = response.headers.get('content-type', '')
            content_disp = response.headers.get('content-disposition', '')

            print(f"\n  Endpoint: {endpoint[:60]}...")
            print(f"  Status: {response.status_code}")
            print(f"  Content-Type: {content_type}")
            print(f"  Content-Disposition: {content_disp}")

            if response.status_code == 200 and ('zip' in content_type or 'pdf' in content_type or 'octet' in content_type):
                print(f"  [OK] Download disponivel!")
                return endpoint

        except Exception as e:
            print(f"  [X] Erro: {e}")

    return None


def obter_info_fundo(token: str, uuid: str):
    """Obtem informacoes do fundo."""
    print(f"\n[4] INFORMACOES DO FUNDO")
    print("=" * 50)

    url = f'https://hub.qoredtvm.com.br/api/v1/fundos-posicao/{uuid}'

    headers = {
        'Authorization': f'Bearer {token}',
        'Accept': 'application/json'
    }

    try:
        response = requests.get(url, headers=headers, timeout=10)

        if response.status_code == 200:
            data = response.json()
            print(f"[+] Nome: {data.get('nome', 'N/A')}")
            print(f"[+] CNPJ: {data.get('cnpj', 'N/A')}")
            print(f"[+] Chaves: {list(data.keys())}")
            return data

    except Exception as e:
        print(f"[X] Erro: {e}")

    return None


def buscar_endpoints_xml(token: str, uuid: str):
    """Busca endpoints especificos para XML."""
    print(f"\n[6] BUSCA DE ENDPOINTS XML")
    print("=" * 50)

    headers = {
        'Authorization': f'Bearer {token}',
        'Accept': 'application/json, application/xml, */*'
    }

    # Endpoints possiveis para XML
    endpoints = [
        f'/api/v1/fundos-posicao/{uuid}/xml',
        f'/api/v1/fundos-posicao/{uuid}/xml-anbima',
        f'/api/v1/fundos-posicao/{uuid}/anbima',
        f'/api/v1/fundos-posicao/{uuid}/relatorios',
        f'/api/v1/fundos-posicao/{uuid}/relatorios/xml',
        f'/api/v1/fundos-posicao/{uuid}/relatorios/anbima',
        f'/api/v1/fundos-posicao/{uuid}/exportar',
        f'/api/v1/fundos-posicao/{uuid}/export',
        f'/api/v1/fundos-posicao/{uuid}/download',
        f'/api/v1/fundos/{uuid}/xml',
        f'/api/v1/fundos/{uuid}/arquivos',
        f'/api/v1/xml/{uuid}',
        f'/api/v1/anbima/{uuid}',
        f'/api/v1/relatorios/{uuid}',
        f'/api/v1/fundos-posicao/{uuid}/carteira-xml',
        f'/api/v1/fundos-posicao/{uuid}/posicao-xml',
        # Com query params
        f'/api/v1/fundos-posicao/{uuid}/arquivos?formato=xml',
        f'/api/v1/fundos-posicao/{uuid}/arquivos?type=xml',
        f'/api/v1/fundos-posicao/{uuid}/arquivos?output=xml',
    ]

    print(f"[+] Testando {len(endpoints)} endpoints possiveis...")

    encontrados = []

    for endpoint in endpoints:
        url = f'https://hub.qoredtvm.com.br{endpoint}'

        try:
            response = requests.get(url, headers=headers, timeout=10)
            content_type = response.headers.get('content-type', '')

            status_ok = response.status_code in [200, 201]
            has_content = 'xml' in content_type or 'json' in content_type

            if status_ok:
                print(f"  [OK] {endpoint}")
                print(f"       Status: {response.status_code}, Type: {content_type[:50]}")
                encontrados.append({
                    'endpoint': endpoint,
                    'status': response.status_code,
                    'content_type': content_type
                })

                # Mostra preview da resposta
                try:
                    if 'json' in content_type:
                        data = response.json()
                        if isinstance(data, dict):
                            print(f"       Chaves: {list(data.keys())[:5]}")
                        elif isinstance(data, list):
                            print(f"       Lista com {len(data)} itens")
                except:
                    pass
            elif response.status_code == 400:
                # 400 pode indicar parametro faltando
                print(f"  [??] {endpoint} -> 400 (pode precisar de params)")
            else:
                print(f"  [--] {endpoint} -> {response.status_code}")

        except Exception as e:
            print(f"  [X] {endpoint} -> {str(e)[:50]}")

    if encontrados:
        print(f"\n[+] ENDPOINTS ENCONTRADOS: {len(encontrados)}")
        for e in encontrados:
            print(f"    - {e['endpoint']}")
    else:
        print("\n[!] Nenhum endpoint XML encontrado")

    return encontrados


def main():
    print("\n" + "=" * 70)
    print("  TESTE DE API QORE - DESCOBERTA DE ENDPOINTS")
    print("=" * 70)

    # Carrega credenciais
    PLANILHA = r'C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\03. Arquivos Rotina\07. DEPARA\DOWNLOADS_AUX.xlsx'

    credentials = carregar_credenciais(PLANILHA)
    if not credentials:
        return

    # Autentica
    token = autenticar(credentials)
    if not token:
        print("[X] Falha na autenticacao")
        return

    # Carrega lista de fundos para pegar um UUID de teste
    fundos_path = Path(__file__).parent.parent / 'dump_qore' / 'fundos' / 'lista_fundos.json'

    if fundos_path.exists():
        with open(fundos_path, 'r', encoding='utf-8') as f:
            fundos = json.load(f)

        if fundos:
            url_fundo = fundos[0]['url']
            uuid = url_fundo.split('/fundo-posicao/')[-1].split('/')[0]
            nome_fundo = fundos[0]['nome']

            print(f"\n[+] Usando fundo de teste: {nome_fundo}")
            print(f"[+] UUID: {uuid}")

            # Obtem info do fundo
            obter_info_fundo(token, uuid)

            # Testa tipos de arquivo
            tipos_validos = testar_tipos_arquivo(token, uuid)

            # Analisa estrutura dos arquivos
            if tipos_validos:
                arquivo_info = analisar_estrutura_arquivos(token, uuid, tipos_validos[0]['tipo'])

                # Testa download individual
                if arquivo_info:
                    testar_download_individual(token, uuid, arquivo_info)

                # Testa download em lote
                testar_download(token, uuid, tipos_validos[0]['tipo'])

            # Busca endpoints XML especificos
            buscar_endpoints_xml(token, uuid)
    else:
        print(f"[X] Arquivo de fundos nao encontrado: {fundos_path}")
        print("[!] Execute primeiro o captura_qore_cdp.py para gerar a lista")

    print("\n" + "=" * 70)
    print("  TESTE CONCLUIDO")
    print("=" * 70 + "\n")


if __name__ == '__main__':
    main()
