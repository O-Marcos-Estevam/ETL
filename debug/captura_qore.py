"""
================================================================================
CAPTURA QORE - Script de Diagnostico e Mapeamento
================================================================================

Este script faz login no QORE e captura:
- HTML de todas as paginas relevantes
- Chamadas de API (endpoints, headers, payloads)
- Cookies e tokens de autenticacao
- Estrutura de navegacao do site
- Scripts JS carregados
- Informacoes dos fundos disponiveis

Uso:
    python captura_qore.py

Requisitos:
    pip install selenium-wire beautifulsoup4 openpyxl

Autor: ETL Team
Data: Dezembro 2025
================================================================================
"""

import os
import sys
import json
import time
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
from urllib.parse import urljoin, urlparse
from dataclasses import dataclass, asdict
from collections import defaultdict

# Adiciona path do projeto
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from seleniumwire import webdriver
    SELENIUM_WIRE_AVAILABLE = True
except ImportError:
    from selenium import webdriver
    SELENIUM_WIRE_AVAILABLE = False
    print("[!] selenium-wire nao instalado. Instale com: pip install selenium-wire")
    print("[!] Continuando com selenium padrao (sem captura de rede)")

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

try:
    from bs4 import BeautifulSoup
    BS4_AVAILABLE = True
except ImportError:
    BS4_AVAILABLE = False
    print("[!] beautifulsoup4 nao instalado. Instale com: pip install beautifulsoup4")

import openpyxl


# =============================================================================
# CONFIGURACAO
# =============================================================================

@dataclass
class CapturaConfig:
    """Configuracoes da captura."""
    url_base: str = "https://hub.qoredtvm.com.br"
    output_dir: str = "dump_qore"
    max_fundos_captura: int = 5  # Quantos fundos navegar para capturar endpoints
    timeout_pagina: int = 15
    delay_entre_paginas: float = 1.5
    capturar_js: bool = True
    capturar_css: bool = False
    verbose: bool = True


# =============================================================================
# CLASSE PRINCIPAL
# =============================================================================

class CapturaQore:
    """Captura informacoes completas do portal QORE."""

    def __init__(self, config: CapturaConfig = None):
        self.config = config or CapturaConfig()
        self.driver = None
        self.output_path = Path(self.config.output_dir)

        # Dados capturados
        self.api_calls: List[Dict] = []
        self.paginas_html: Dict[str, str] = {}
        self.fundos_encontrados: List[Dict] = []
        self.cookies: List[Dict] = []
        self.scripts_js: List[str] = []
        self.endpoints_unicos: Dict[str, Dict] = {}
        self.headers_auth: Dict[str, str] = {}

        # Estatisticas
        self.stats = {
            'paginas_capturadas': 0,
            'api_calls': 0,
            'fundos_encontrados': 0,
            'endpoints_unicos': 0,
            'erros': 0
        }

    def _log(self, msg: str, level: str = "INFO"):
        """Log com timestamp."""
        if self.config.verbose or level in ["ERROR", "CRITICAL"]:
            symbols = {"INFO": "[+]", "WARNING": "[!]", "ERROR": "[X]", "DEBUG": "[.]"}
            symbol = symbols.get(level, "[?]")
            print(f"{symbol} {msg}")

    def _criar_estrutura_pastas(self):
        """Cria estrutura de pastas para output."""
        pastas = [
            self.output_path,
            self.output_path / "html",
            self.output_path / "api",
            self.output_path / "js",
            self.output_path / "screenshots",
            self.output_path / "fundos"
        ]
        for pasta in pastas:
            pasta.mkdir(parents=True, exist_ok=True)
        self._log(f"Estrutura criada em: {self.output_path}")

    def _iniciar_driver(self):
        """Inicia Chrome com interceptacao de rede."""
        self._log("Iniciando Chrome Driver...")

        chrome_options = Options()

        # Configuracoes basicas
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--start-maximized')

        # Performance (mas mantendo recursos para captura)
        chrome_options.add_argument('--disable-extensions')
        chrome_options.add_argument('--disable-notifications')
        chrome_options.add_argument('--log-level=3')

        # Preferencias
        prefs = {
            'download.prompt_for_download': False,
            'profile.default_content_settings.popups': 0
        }
        chrome_options.add_experimental_option('prefs', prefs)

        if SELENIUM_WIRE_AVAILABLE:
            # Configuracao do selenium-wire para capturar requests
            seleniumwire_options = {
                'disable_encoding': True,  # Descomprime respostas
                'suppress_connection_errors': True
            }
            self.driver = webdriver.Chrome(
                options=chrome_options,
                seleniumwire_options=seleniumwire_options
            )
        else:
            self.driver = webdriver.Chrome(options=chrome_options)

        self.driver.implicitly_wait(5)
        self._log("Chrome Driver iniciado com sucesso")

    def _fazer_login(self, email: str, senha: str) -> bool:
        """Realiza login no QORE."""
        try:
            self._log(f"Acessando: {self.config.url_base}")
            self.driver.get(self.config.url_base)
            time.sleep(2)

            # Salva pagina de login
            self._salvar_html("01_login", self.driver.page_source)
            self._screenshot("01_login")

            # Preenche credenciais
            self._log("Inserindo credenciais...")

            email_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.NAME, 'email'))
            )
            email_field.clear()
            email_field.send_keys(email)

            senha_field = self.driver.find_element(By.NAME, 'password')
            senha_field.clear()
            senha_field.send_keys(senha + Keys.RETURN)

            # Aguarda dashboard
            WebDriverWait(self.driver, self.config.timeout_pagina).until(
                lambda d: 'dashboard' in d.current_url.lower()
            )

            self._log("Login realizado com sucesso!")
            time.sleep(2)

            # Captura cookies apos login
            self._capturar_cookies()

            return True

        except Exception as e:
            self._log(f"Falha no login: {e}", "ERROR")
            self._screenshot("erro_login")
            return False

    def _capturar_cookies(self):
        """Captura e salva cookies da sessao."""
        self.cookies = self.driver.get_cookies()

        # Salva cookies
        cookies_file = self.output_path / "cookies.json"
        with open(cookies_file, 'w', encoding='utf-8') as f:
            json.dump(self.cookies, f, indent=2, ensure_ascii=False)

        self._log(f"Capturados {len(self.cookies)} cookies")

        # Extrai tokens de autenticacao
        for cookie in self.cookies:
            name = cookie.get('name', '').lower()
            if any(x in name for x in ['token', 'auth', 'session', 'jwt']):
                self.headers_auth[cookie['name']] = cookie['value']
                self._log(f"  Token encontrado: {cookie['name']}")

    def _capturar_requests(self):
        """Captura todas as requisicoes de rede (selenium-wire)."""
        if not SELENIUM_WIRE_AVAILABLE:
            return

        for req in self.driver.requests:
            if not req.response:
                continue

            url = req.url
            method = req.method
            status = req.response.status_code
            content_type = req.response.headers.get('Content-Type', '')

            # Filtra assets estaticos
            is_asset = any(ext in url.lower() for ext in [
                '.js', '.css', '.png', '.jpg', '.jpeg', '.gif',
                '.svg', '.woff', '.woff2', '.ttf', '.ico', '.map'
            ])

            # Identifica chamadas de API
            is_api = any(x in url.lower() for x in [
                '/api/', 'graphql', '/v1/', '/v2/',
                'fundos', 'boletas', 'carteira', 'download',
                'relatorio', 'posicao', 'cota', 'patrimonio'
            ]) or 'json' in content_type

            if is_api or (not is_asset and status == 200):
                # Extrai headers de autorizacao
                auth_header = None
                if req.headers:
                    auth_header = req.headers.get('Authorization')
                    if auth_header and 'Authorization' not in self.headers_auth:
                        self.headers_auth['Authorization'] = auth_header

                call_info = {
                    'method': method,
                    'url': url,
                    'status': status,
                    'content_type': content_type,
                    'response_size': len(req.response.body) if req.response.body else 0,
                    'auth_header': auth_header,
                    'request_headers': dict(req.headers) if req.headers else {},
                    'timestamp': datetime.now().isoformat()
                }

                # Tenta capturar body da resposta (JSON)
                if 'json' in content_type and req.response.body:
                    try:
                        call_info['response_body'] = json.loads(req.response.body.decode('utf-8'))
                    except:
                        pass

                self.api_calls.append(call_info)

                # Endpoint unico
                parsed = urlparse(url)
                endpoint_key = f"{method} {parsed.path}"
                if endpoint_key not in self.endpoints_unicos:
                    self.endpoints_unicos[endpoint_key] = call_info
                    self._log(f"  API: [{method}] {parsed.path}", "DEBUG")

        # Limpa requests para proxima captura
        del self.driver.requests

    def _salvar_html(self, nome: str, html: str):
        """Salva HTML de uma pagina."""
        arquivo = self.output_path / "html" / f"{nome}.html"
        arquivo.write_text(html, encoding='utf-8')
        self.paginas_html[nome] = html
        self.stats['paginas_capturadas'] += 1

    def _screenshot(self, nome: str):
        """Salva screenshot."""
        arquivo = self.output_path / "screenshots" / f"{nome}.png"
        try:
            self.driver.save_screenshot(str(arquivo))
        except:
            pass

    def _extrair_scripts_js(self, html: str) -> List[str]:
        """Extrai URLs de scripts JS do HTML."""
        if not BS4_AVAILABLE:
            # Fallback com regex
            pattern = r'<script[^>]+src=["\']([^"\']+)["\']'
            return re.findall(pattern, html)

        soup = BeautifulSoup(html, 'html.parser')
        scripts = []

        for script in soup.find_all('script', src=True):
            src = script.get('src')
            if src:
                scripts.append(src)

        return scripts

    def _capturar_dashboard(self):
        """Captura informacoes do dashboard."""
        self._log("Capturando dashboard...")

        html = self.driver.page_source
        self._salvar_html("02_dashboard", html)
        self._screenshot("02_dashboard")

        # Extrai scripts JS
        scripts = self._extrair_scripts_js(html)
        self.scripts_js.extend(scripts)
        self._log(f"  {len(scripts)} scripts JS encontrados")

        # Captura requests ate agora
        self._capturar_requests()

        # Extrai links de fundos
        self._extrair_fundos_dashboard()

    def _extrair_fundos_dashboard(self):
        """Extrai lista de fundos do dashboard."""
        self._log("Extraindo lista de fundos...")

        try:
            # Busca links que parecem ser fundos
            links = self.driver.find_elements(By.XPATH, "//a[contains(@href, 'fundos') or contains(@href, 'fundo')]")

            for link in links:
                try:
                    href = link.get_attribute('href')
                    texto = link.text.strip()

                    if href and texto:
                        self.fundos_encontrados.append({
                            'nome': texto,
                            'url': href,
                            'capturado': False
                        })
                except:
                    continue

            # Remove duplicatas
            urls_vistas = set()
            fundos_unicos = []
            for f in self.fundos_encontrados:
                if f['url'] not in urls_vistas:
                    urls_vistas.add(f['url'])
                    fundos_unicos.append(f)

            self.fundos_encontrados = fundos_unicos
            self.stats['fundos_encontrados'] = len(self.fundos_encontrados)

            self._log(f"  {len(self.fundos_encontrados)} fundos encontrados")

            # Salva lista de fundos
            fundos_file = self.output_path / "fundos" / "lista_fundos.json"
            with open(fundos_file, 'w', encoding='utf-8') as f:
                json.dump(self.fundos_encontrados, f, indent=2, ensure_ascii=False)

        except Exception as e:
            self._log(f"Erro ao extrair fundos: {e}", "ERROR")

    def _capturar_pagina_fundo(self, fundo: Dict, indice: int):
        """Captura pagina de um fundo especifico."""
        nome = fundo['nome']
        url = fundo['url']

        self._log(f"Capturando fundo {indice + 1}: {nome}")

        try:
            self.driver.get(url)
            time.sleep(self.config.delay_entre_paginas)

            # Salva HTML
            html = self.driver.page_source
            nome_arquivo = f"fundo_{indice + 1:02d}_{self._sanitizar_nome(nome)}"
            self._salvar_html(nome_arquivo, html)
            self._screenshot(nome_arquivo)

            # Captura requests
            self._capturar_requests()

            # Tenta encontrar botoes de relatorio
            self._mapear_botoes_relatorio(nome)

            fundo['capturado'] = True

        except Exception as e:
            self._log(f"Erro ao capturar {nome}: {e}", "ERROR")
            self.stats['erros'] += 1

    def _mapear_botoes_relatorio(self, nome_fundo: str):
        """Mapeia botoes de relatorio disponveis na pagina do fundo."""
        botoes_info = []

        try:
            # Busca botoes com texto de relatorio
            botoes = self.driver.find_elements(By.XPATH,
                "//button[contains(., 'Carteira') or contains(., 'PDF') or contains(., 'Excel') or contains(., 'XML') or contains(., 'Download')]"
            )

            for btn in botoes:
                try:
                    texto = btn.text.strip()
                    classe = btn.get_attribute('class')
                    onclick = btn.get_attribute('onclick')

                    if texto:
                        botoes_info.append({
                            'texto': texto,
                            'classe': classe,
                            'onclick': onclick
                        })
                except:
                    continue

            if botoes_info:
                self._log(f"  Botoes encontrados: {[b['texto'] for b in botoes_info]}", "DEBUG")

                # Tenta clicar em um botao para capturar endpoint
                for btn_info in botoes_info:
                    if 'XML' in btn_info['texto'] or 'Excel' in btn_info['texto']:
                        try:
                            btn = self.driver.find_element(By.XPATH, f"//button[contains(., '{btn_info['texto']}')]")
                            btn.click()
                            time.sleep(1)
                            self._capturar_requests()

                            # Tenta encontrar menu de download em lote
                            self._capturar_menu_download_lote()

                            # Volta
                            self.driver.back()
                            time.sleep(0.5)
                            break
                        except:
                            pass

        except Exception as e:
            self._log(f"  Erro ao mapear botoes: {e}", "DEBUG")

    def _capturar_menu_download_lote(self):
        """Captura informacoes do menu de download em lote."""
        try:
            # Procura menu de opcoes (...)
            menus = self.driver.find_elements(By.XPATH, '//div[@data-kt-menu-trigger="click"]')

            for menu in menus:
                try:
                    icone = menu.find_element(By.TAG_NAME, 'i')
                    if 'ellipsis' in icone.get_attribute('class'):
                        menu.click()
                        time.sleep(0.5)

                        # Captura HTML do menu aberto
                        self._salvar_html("menu_download_lote", self.driver.page_source)
                        self._capturar_requests()

                        # Clica em Download em Lote
                        try:
                            link_lote = self.driver.find_element(By.XPATH, '//a[contains(., "Download em Lote")]')
                            link_lote.click()
                            time.sleep(0.5)

                            # Captura modal de download
                            self._salvar_html("modal_download_lote", self.driver.page_source)
                            self._screenshot("modal_download_lote")
                            self._capturar_requests()

                            # Fecha modal (ESC)
                            from selenium.webdriver.common.keys import Keys
                            self.driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)

                        except:
                            pass

                        break
                except:
                    continue

        except Exception as e:
            self._log(f"  Erro ao capturar menu: {e}", "DEBUG")

    def _sanitizar_nome(self, nome: str) -> str:
        """Sanitiza nome para usar em arquivo."""
        return re.sub(r'[^\w\-]', '_', nome)[:50]

    def _baixar_scripts_js(self):
        """Baixa scripts JS para analise."""
        if not self.config.capturar_js or not self.scripts_js:
            return

        self._log("Baixando scripts JS...")

        import requests
        session = requests.Session()

        # Copia cookies
        for cookie in self.cookies:
            session.cookies.set(cookie['name'], cookie['value'])

        base_url = self.config.url_base
        js_dir = self.output_path / "js"

        for script_url in self.scripts_js[:20]:  # Limita a 20 scripts
            try:
                # Resolve URL relativa
                if script_url.startswith('//'):
                    script_url = 'https:' + script_url
                elif script_url.startswith('/'):
                    script_url = urljoin(base_url, script_url)
                elif not script_url.startswith('http'):
                    script_url = urljoin(base_url, script_url)

                # Baixa
                response = session.get(script_url, timeout=10)
                if response.ok:
                    nome_arquivo = script_url.split('/')[-1].split('?')[0]
                    if not nome_arquivo.endswith('.js'):
                        nome_arquivo += '.js'

                    arquivo = js_dir / nome_arquivo
                    arquivo.write_bytes(response.content)
                    self._log(f"  JS baixado: {nome_arquivo}", "DEBUG")

            except Exception as e:
                self._log(f"  Erro ao baixar JS: {e}", "DEBUG")

    def _analisar_scripts_js(self):
        """Analisa scripts JS para encontrar endpoints."""
        self._log("Analisando scripts JS...")

        js_dir = self.output_path / "js"
        endpoints_encontrados = []

        # Padroes para buscar endpoints
        patterns = [
            r'["\']/(api/[^"\']+)["\']',
            r'["\']/(v\d+/[^"\']+)["\']',
            r'fetch\s*\(\s*["\']([^"\']+)["\']',
            r'axios\.[a-z]+\s*\(\s*["\']([^"\']+)["\']',
            r'url:\s*["\']([^"\']+)["\']',
            r'endpoint:\s*["\']([^"\']+)["\']',
        ]

        for js_file in js_dir.glob('*.js'):
            try:
                content = js_file.read_text(encoding='utf-8', errors='ignore')

                for pattern in patterns:
                    matches = re.findall(pattern, content)
                    for match in matches:
                        if match and '/' in match and not match.startswith('//'):
                            endpoints_encontrados.append({
                                'endpoint': match,
                                'arquivo': js_file.name
                            })

            except Exception as e:
                continue

        # Remove duplicatas
        endpoints_unicos = []
        vistos = set()
        for ep in endpoints_encontrados:
            if ep['endpoint'] not in vistos:
                vistos.add(ep['endpoint'])
                endpoints_unicos.append(ep)

        if endpoints_unicos:
            self._log(f"  {len(endpoints_unicos)} endpoints encontrados em JS")

            # Salva
            arquivo = self.output_path / "api" / "endpoints_js.json"
            with open(arquivo, 'w', encoding='utf-8') as f:
                json.dump(endpoints_unicos, f, indent=2, ensure_ascii=False)

    def _gerar_relatorio(self):
        """Gera relatorio final da captura."""
        self._log("Gerando relatorio...")

        relatorio = {
            'data_captura': datetime.now().isoformat(),
            'url_base': self.config.url_base,
            'estatisticas': self.stats,
            'endpoints_api': list(self.endpoints_unicos.keys()),
            'headers_autenticacao': {k: v[:20] + '...' if len(v) > 20 else v
                                      for k, v in self.headers_auth.items()},
            'fundos_capturados': len([f for f in self.fundos_encontrados if f.get('capturado')]),
            'total_fundos': len(self.fundos_encontrados),
            'total_api_calls': len(self.api_calls),
            'scripts_js': len(self.scripts_js)
        }

        # Salva relatorio
        relatorio_file = self.output_path / "RELATORIO.json"
        with open(relatorio_file, 'w', encoding='utf-8') as f:
            json.dump(relatorio, f, indent=2, ensure_ascii=False)

        # Salva API calls completas
        api_file = self.output_path / "api" / "todas_chamadas.json"
        with open(api_file, 'w', encoding='utf-8') as f:
            json.dump(self.api_calls, f, indent=2, ensure_ascii=False, default=str)

        # Salva endpoints unicos
        endpoints_file = self.output_path / "api" / "endpoints_unicos.json"
        with open(endpoints_file, 'w', encoding='utf-8') as f:
            json.dump(self.endpoints_unicos, f, indent=2, ensure_ascii=False, default=str)

        # Salva headers de auth
        auth_file = self.output_path / "auth_headers.json"
        with open(auth_file, 'w', encoding='utf-8') as f:
            json.dump(self.headers_auth, f, indent=2, ensure_ascii=False)

        self.stats['endpoints_unicos'] = len(self.endpoints_unicos)
        self.stats['api_calls'] = len(self.api_calls)

        # Imprime resumo
        print("\n" + "=" * 70)
        print("  RELATORIO DE CAPTURA")
        print("=" * 70)
        print(f"  Paginas HTML capturadas:  {self.stats['paginas_capturadas']}")
        print(f"  Fundos encontrados:       {self.stats['fundos_encontrados']}")
        print(f"  Chamadas de API:          {self.stats['api_calls']}")
        print(f"  Endpoints unicos:         {self.stats['endpoints_unicos']}")
        print(f"  Scripts JS:               {len(self.scripts_js)}")
        print(f"  Erros:                    {self.stats['erros']}")
        print("=" * 70)
        print(f"  Output salvo em: {self.output_path.absolute()}")
        print("=" * 70)

        if self.endpoints_unicos:
            print("\n  ENDPOINTS ENCONTRADOS:")
            for ep in list(self.endpoints_unicos.keys())[:15]:
                print(f"    - {ep}")
            if len(self.endpoints_unicos) > 15:
                print(f"    ... e mais {len(self.endpoints_unicos) - 15}")

    def executar(self, email: str, senha: str):
        """Executa captura completa."""
        print("\n" + "=" * 70)
        print("  CAPTURA QORE - Diagnostico Completo")
        print("=" * 70 + "\n")

        self._criar_estrutura_pastas()

        try:
            self._iniciar_driver()

            if not self._fazer_login(email, senha):
                return

            self._capturar_dashboard()

            # Captura alguns fundos para mapear endpoints
            fundos_para_capturar = self.fundos_encontrados[:self.config.max_fundos_captura]

            for i, fundo in enumerate(fundos_para_capturar):
                self._capturar_pagina_fundo(fundo, i)

            # Baixa e analisa JS
            self._baixar_scripts_js()
            self._analisar_scripts_js()

            # Gera relatorio
            self._gerar_relatorio()

        except Exception as e:
            self._log(f"Erro fatal: {e}", "CRITICAL")
            import traceback
            traceback.print_exc()

        finally:
            if self.driver:
                self.driver.quit()
                self._log("Chrome fechado")


# =============================================================================
# FUNCAO PARA CARREGAR CREDENCIAIS DA PLANILHA
# =============================================================================

def carregar_credenciais_planilha(caminho: str) -> tuple:
    """Carrega credenciais da planilha DOWNLOADS_AUX.xlsx."""
    try:
        wb = openpyxl.load_workbook(caminho, data_only=True)
        ws = wb['Downloads']

        url = str(ws['M10'].value or '').strip()
        email = str(ws['N10'].value or '').strip()
        senha = str(ws['O10'].value or '').strip()

        return url, email, senha

    except Exception as e:
        print(f"[X] Erro ao ler planilha: {e}")
        return None, None, None


# =============================================================================
# MAIN
# =============================================================================

if __name__ == '__main__':
    # Caminho padrao da planilha de configuracao
    PLANILHA_PATH = r'C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\03. Arquivos Rotina\07. DEPARA\DOWNLOADS_AUX.xlsx'

    # Tenta carregar credenciais da planilha
    url, email, senha = carregar_credenciais_planilha(PLANILHA_PATH)

    if not all([url, email, senha]):
        print("\n[!] Credenciais nao encontradas na planilha.")
        print("[!] Por favor, edite este arquivo e configure manualmente:\n")

        url = input("URL do QORE (ex: https://hub.qoredtvm.com.br): ").strip()
        email = input("Email: ").strip()
        senha = input("Senha: ").strip()

    if not all([url, email, senha]):
        print("\n[X] Credenciais invalidas. Abortando.")
        sys.exit(1)

    # Configura e executa
    config = CapturaConfig(
        url_base=url,
        output_dir="dump_qore",
        max_fundos_captura=3,  # Captura 3 fundos para mapear endpoints
        verbose=True
    )

    captura = CapturaQore(config)
    captura.executar(email, senha)
