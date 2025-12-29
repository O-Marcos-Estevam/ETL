"""
================================================================================
CAPTURA ENDPOINT XML - Descobre o endpoint usado para XML
================================================================================

Clica especificamente no botao XML Anbima 5.0 e captura as chamadas de rede.

Uso: python debug/captura_xml_endpoint.py
================================================================================
"""

import os
import sys
import json
import time
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl


def carregar_credenciais(planilha_path: str):
    """Carrega credenciais da planilha."""
    try:
        wb = openpyxl.load_workbook(planilha_path, data_only=True)
        ws = wb['Downloads']
        return {
            'url': str(ws['M10'].value or '').strip(),
            'email': str(ws['N10'].value or '').strip(),
            'senha': str(ws['O10'].value or '').strip()
        }
    except Exception as e:
        print(f"[X] Erro ao carregar credenciais: {e}")
        return None


def extrair_logs_rede(driver) -> list:
    """Extrai logs de rede do Chrome."""
    logs = []
    try:
        perf_logs = driver.get_log('performance')
        for entry in perf_logs:
            try:
                log = json.loads(entry['message'])
                message = log.get('message', {})
                method = message.get('method', '')
                params = message.get('params', {})

                if method == 'Network.requestWillBeSent':
                    request = params.get('request', {})
                    url = request.get('url', '')
                    logs.append({
                        'type': 'request',
                        'method': request.get('method'),
                        'url': url,
                        'postData': request.get('postData'),
                    })
                elif method == 'Network.responseReceived':
                    response = params.get('response', {})
                    logs.append({
                        'type': 'response',
                        'url': response.get('url', ''),
                        'status': response.get('status'),
                        'mimeType': response.get('mimeType'),
                    })
            except:
                continue
    except:
        pass
    return logs


def main():
    print("\n" + "=" * 70)
    print("  CAPTURA ENDPOINT XML - Chrome DevTools Protocol")
    print("=" * 70)

    PLANILHA = r'C:\bloko\Fundos - Documentos\00. Monitoramento\01. Rotinas\03. Arquivos Rotina\07. DEPARA\DOWNLOADS_AUX.xlsx'

    credentials = carregar_credenciais(PLANILHA)
    if not credentials:
        return

    # Carrega lista de fundos
    fundos_path = Path(__file__).parent.parent / 'dump_qore' / 'fundos' / 'lista_fundos.json'
    with open(fundos_path, 'r', encoding='utf-8') as f:
        fundos = json.load(f)

    fundo_url = fundos[0]['url']
    fundo_nome = fundos[0]['nome']

    print(f"\n[+] Fundo: {fundo_nome}")
    print(f"[+] URL: {fundo_url}")

    # Configura Chrome com logging
    chrome_options = Options()
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.set_capability('goog:loggingPrefs', {'performance': 'ALL'})

    print("\n[+] Iniciando Chrome...")
    driver = webdriver.Chrome(options=chrome_options)
    driver.execute_cdp_cmd('Network.enable', {})

    try:
        # Login
        print("[+] Fazendo login...")
        driver.get(credentials['url'])
        time.sleep(2)

        email_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, 'email'))
        )
        email_field.send_keys(credentials['email'])

        senha_field = driver.find_element(By.NAME, 'password')
        senha_field.send_keys(credentials['senha'] + Keys.RETURN)

        WebDriverWait(driver, 15).until(
            lambda d: 'dashboard' in d.current_url.lower()
        )
        print("[+] Login OK!")
        time.sleep(2)

        # Navega para o fundo
        print(f"[+] Navegando para fundo...")
        driver.get(fundo_url)
        time.sleep(3)

        # Limpa logs anteriores
        extrair_logs_rede(driver)

        # Procura e clica no botao XML Anbima 5.0
        print("[+] Procurando botao XML Anbima 5.0...")

        botoes = driver.find_elements(By.XPATH, "//button")
        xml_button = None

        for btn in botoes:
            texto = btn.text.strip()
            print(f"    Botao encontrado: '{texto}'")
            if 'XML' in texto.upper() and '5.0' in texto:
                xml_button = btn
                print(f"    [!] Botao XML encontrado: '{texto}'")
                break

        if xml_button:
            print("\n[+] Clicando no botao XML Anbima 5.0...")
            xml_button.click()
            time.sleep(3)

            # Captura logs apos o clique
            print("\n[+] Capturando chamadas de rede...")
            logs = extrair_logs_rede(driver)

            # Filtra requests relevantes
            print("\n" + "=" * 70)
            print("  CHAMADAS DE REDE APOS CLIQUE NO XML")
            print("=" * 70)

            for log in logs:
                url = log.get('url', '')
                # Ignora assets estaticos
                if any(x in url.lower() for x in ['.js', '.css', '.png', '.jpg', '.svg', '.woff']):
                    continue

                # Mostra requests interessantes
                if 'api' in url.lower() or 'xml' in url.lower() or 'anbima' in url.lower() or 'arquivo' in url.lower():
                    print(f"\n  [{log['type'].upper()}]")
                    print(f"    Method: {log.get('method', 'N/A')}")
                    print(f"    URL: {url}")
                    if log.get('postData'):
                        print(f"    Body: {log.get('postData')[:200]}")
                    if log.get('status'):
                        print(f"    Status: {log.get('status')}")
                    if log.get('mimeType'):
                        print(f"    Type: {log.get('mimeType')}")

            # Verifica se abriu nova aba/modal
            print("\n[+] Verificando janelas/abas abertas...")
            handles = driver.window_handles
            print(f"    {len(handles)} janela(s) aberta(s)")

            # Verifica URL atual
            print(f"\n[+] URL atual: {driver.current_url}")

            # Salva todos os logs para analise
            output_path = Path(__file__).parent.parent / 'dump_qore' / 'api' / 'xml_network_logs.json'
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(logs, f, indent=2, ensure_ascii=False)
            print(f"\n[+] Logs salvos em: {output_path}")

        else:
            print("[X] Botao XML Anbima 5.0 nao encontrado!")
            print("[!] Botoes disponiveis na pagina:")
            for btn in botoes:
                print(f"    - {btn.text.strip()}")

    except Exception as e:
        print(f"[X] Erro: {e}")
        import traceback
        traceback.print_exc()

    finally:
        input("\n[!] Pressione ENTER para fechar o navegador...")
        driver.quit()

    print("\n" + "=" * 70)
    print("  CAPTURA CONCLUIDA")
    print("=" * 70 + "\n")


if __name__ == '__main__':
    main()
